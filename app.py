import streamlit as st
import json, math, os, requests, uuid, io, re, html as _html
from datetime import datetime
import streamlit.components.v1 as components

def esc(s):
    """HTML-escape a string so it's safe to embed in unsafe_allow_html blocks."""
    return _html.escape(str(s) if s is not None else "")

st.set_page_config(page_title="FTA Reverse Engineer", page_icon="⚠️",
                   layout="wide", initial_sidebar_state="expanded")

# ── Constants ─────────────────────────────────────────────────────────────
LEVEL_ORDER        = ["HAZARD", "SF", "FF", "IF", "GROUP"]
LEVEL_COLORS       = {"HAZARD":"#ff4d4d","SF":"#ff8c42","FF":"#f5c518","IF":"#4caf7d","GROUP":"#7e57c2"}
LEVEL_TEXT         = {"HAZARD":"#fff","SF":"#fff","FF":"#111","IF":"#fff","GROUP":"#fff"}
VALID_PARENT_TYPES = ["HAZARD","SF","FF","GROUP"]
VALID_CHILD_TYPES  = ["SF","FF","IF","GROUP"]
DISPLAY_ORDER      = ["HAZARD","SF","FF","IF","GROUP"]

# ── PFTA Forward Verification Engine ─────────────────────────────────────
def builtin_forward_verify(nodes):
    """
    Built-in forward verification engine — no external dependencies.

    Performs a bottom-up forward pass through the fault tree using the
    calculated leaf values, then compares the forward-computed hazard
    probabilities against the stored targetValues.

    For OR gates:  P(top) ≈ sum of children (rare-event approximation)
    For AND gates: P(top) = product of children

    Shared events (same nodeId) are deduplicated — counted only once per gate.

    Returns:
        {
          "success": bool,
          "hazard_results": { hazard_name: computed_prob },
          "divergence":     { hazard_name: { target, forward_calc, inverse_calc, delta_pct, ok } },
          "warnings":       [ list of strings ],
          "gate_results":   { node_name: computed_prob }   (all gates for debug)
        }
    """
    if not nodes:
        return {"success": False, "warnings": ["No nodes to verify."],
                "hazard_results": {}, "divergence": {}, "gate_results": {}}

    by_id    = {n["id"]: n for n in nodes}
    warnings = []

    # Build children map
    children_of = {n["id"]: [] for n in nodes}
    for n in nodes:
        for pid in (n.get("parentIds") or []):
            if pid in children_of:
                children_of[pid].append(n["id"])

    # Bottom-up topological order (leaves first, hazards last)
    visited_topo = []
    seen_topo    = set()
    def topo_visit(nid):
        if nid in seen_topo: return
        seen_topo.add(nid)
        for cid in children_of.get(nid, []):
            topo_visit(cid)
        visited_topo.append(nid)
    for n in nodes:
        if n["type"] == "HAZARD":
            topo_visit(n["id"])

    # Forward compute: each node's forward value = combination of its children
    fwd = {}
    for nid in visited_topo:
        n        = by_id[nid]
        children = children_of.get(nid, [])

        if not children:
            # Leaf node — use calculatedValue directly
            v = n.get("calculatedValue")
            fwd[nid] = float(v) if v is not None else 0.0
            continue

        # Collect unique child values (deduplicate shared events by nodeId)
        seen_nids  = set()
        child_vals = []
        for cid in children:
            c      = by_id.get(cid)
            cnid   = (c.get("nodeId") or "").strip() if c else ""
            if cnid and cnid in seen_nids:
                continue  # shared event — count once
            if cnid:
                seen_nids.add(cnid)
            cv = fwd.get(cid, 0.0)
            child_vals.append(cv)

        if not child_vals:
            fwd[nid] = 0.0
            continue

        gate = n.get("gate", "OR")
        if gate == "OR":
            # Rare-event approximation: sum, clamped to 1
            fwd[nid] = min(1.0, sum(child_vals))
        else:  # AND
            p = 1.0
            for v in child_vals:
                p *= v
            fwd[nid] = p

    # Build results
    gate_results  = {by_id[nid]["name"]: fwd[nid] for nid in fwd}
    hazard_results = {}
    divergence     = {}

    for n in nodes:
        if n["type"] != "HAZARD":
            continue
        fwd_prob = fwd.get(n["id"], 0.0)
        hazard_results[n["name"]] = fwd_prob
        target = n.get("targetValue")
        calc   = n.get("calculatedValue")
        if target and target > 0:
            delta_pct = abs(fwd_prob - target) / target * 100
            divergence[n["name"]] = {
                "target":       target,
                "forward_calc": fwd_prob,
                "inverse_calc": calc,
                "delta_pct":    delta_pct,
                "ok":           delta_pct < 5.0
            }

    return {
        "success":        True,
        "hazard_results": hazard_results,
        "divergence":     divergence,
        "warnings":       warnings,
        "gate_results":   gate_results,
    }


def render_pfta_verification(nodes):
    """
    Streamlit UI component for the Forward Verification panel.
    Built-in forward solver — no external dependencies required.
    Runs a bottom-up pass using calculatedValues and compares
    forward-computed hazard probabilities to their targets.
    """
    st.markdown("""
    <div style="background:#0a0f1a;border:1.5px solid #4fc3f744;border-radius:8px;
                padding:10px 14px;margin-bottom:8px;">
      <div style="font-size:9px;color:#4fc3f7;font-weight:700;letter-spacing:2px;margin-bottom:4px;">
        FORWARD VERIFICATION ENGINE
      </div>
      <div style="font-size:9px;color:#666;line-height:1.6;">
        Performs a bottom-up forward pass through the fault tree using your
        current calculated leaf values, then compares the forward-computed
        top-event probabilities against your hazard targets.
        Catches approximation errors from the top-down inverse distribution.
        OR gates use rare-event summation; AND gates use product.
        Shared events (same Node ID) are deduplicated per gate.
      </div>
    </div>""", unsafe_allow_html=True)

    hazards = [n for n in nodes if n["type"] == "HAZARD"]
    if not hazards:
        st.markdown("<div style='color:#555;font-size:11px;'>No HAZARD nodes — nothing to verify.</div>",
                    unsafe_allow_html=True)
        return

    unresolved = [n for n in nodes if n.get("calculatedValue") is None]
    if unresolved:
        st.markdown(
            f"<div style='font-size:9px;color:#f5c518;background:#1a1200;"
            f"border:1px solid #f5c51844;border-radius:5px;padding:6px 10px;'>"
            f"⚠ {len(unresolved)} node(s) have no calculated value. "
            f"Run CALCULATE first before verifying.</div>",
            unsafe_allow_html=True)

    col_btn, col_info = st.columns([2, 3])
    with col_btn:
        run_verify = st.button("🔬 Run Forward Verification",
                               use_container_width=True,
                               key="pfta_verify_btn",
                               help="Bottom-up forward pass — compares computed hazard probabilities to targets")
    with col_info:
        st.markdown(
            "<div style='font-size:9px;color:#555;padding-top:6px;line-height:1.5;'>"
            "Built-in engine — no extra libraries needed. "
            "Rare-event OR summation; AND product. Shared events deduplicated.</div>",
            unsafe_allow_html=True)

    if run_verify:
        with st.spinner("Running forward verification..."):
            result = builtin_forward_verify(nodes)

        if not result["success"]:
            st.error("Verification failed")
            for w in result["warnings"]:
                st.markdown(f"<div style='font-size:10px;color:#ff4d4d;'>{esc(w)}</div>",
                            unsafe_allow_html=True)
            return

        if result["warnings"]:
            for w in result["warnings"]:
                st.markdown(
                    f"<div style='font-size:9px;color:#f5c518;background:#1a1200;"
                    f"border:1px solid #f5c51844;border-radius:4px;padding:4px 8px;margin:2px 0;'>"
                    f"⚠ {esc(w)}</div>", unsafe_allow_html=True)

        # ── Divergence table ─────────────────────────────────────────────
        div = result["divergence"]
        if div:
            all_ok = all(v["ok"] for v in div.values())
            summary_color = "#4caf7d" if all_ok else "#ff4d4d"
            summary_icon  = "✓ ALL WITHIN TOLERANCE (< 5%)" if all_ok else "✗ DIVERGENCE DETECTED"
            st.markdown(
                f"<div style='font-size:10px;color:{summary_color};font-weight:700;"
                f"background:{summary_color}11;border:1px solid {summary_color}44;"
                f"border-radius:5px;padding:6px 12px;margin:6px 0;letter-spacing:1px;'>"
                f"{summary_icon}</div>", unsafe_allow_html=True)

            for hname, d in div.items():
                ok      = d["ok"]
                color   = "#4caf7d" if ok else "#ff4d4d"
                icon    = "✓" if ok else "✗"
                delta   = d["delta_pct"]
                target  = d["target"]
                fwd_p   = d["forward_calc"]
                inv_c   = d["inverse_calc"]
                warn_html = (
                    f'<div style="font-size:9px;color:#ff4d4d;margin-top:5px;line-height:1.4;">'
                    f'⚠ Divergence exceeds 5%. The inverse approximation may be inaccurate '
                    f'for this branch. Check shared events or high-probability paths.</div>'
                ) if not ok else ""

                st.markdown(f"""
                <div style="background:#111;border:1.5px solid {color}44;border-radius:7px;
                            padding:8px 12px;margin:4px 0;">
                  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
                    <span style="font-size:11px;font-weight:700;color:{color};">{icon} {esc(hname)}</span>
                    <span style="font-size:9px;color:{color};background:{'#0a1a0a' if ok else '#1a0a0a'};
                                 padding:1px 8px;border-radius:10px;font-weight:700;">
                      &Delta; {delta:.2f}%
                    </span>
                  </div>
                  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;">
                    <div style="background:#0a0a0a;border-radius:4px;padding:5px 8px;">
                      <div style="font-size:7px;color:#555;letter-spacing:1px;">TARGET</div>
                      <div style="font-size:10px;color:#ff8c42;font-family:monospace;font-weight:700;">{target:.3e}</div>
                    </div>
                    <div style="background:#0a0a0a;border-radius:4px;padding:5px 8px;">
                      <div style="font-size:7px;color:#555;letter-spacing:1px;">INVERSE CALC</div>
                      <div style="font-size:10px;color:#f5c518;font-family:monospace;font-weight:700;">{f"{inv_c:.3e}" if inv_c is not None else "&#8212;"}</div>
                    </div>
                    <div style="background:#0a0a0a;border-radius:4px;padding:5px 8px;">
                      <div style="font-size:7px;color:#555;letter-spacing:1px;">FORWARD COMPUTED</div>
                      <div style="font-size:10px;color:{color};font-family:monospace;font-weight:700;">{fwd_p:.3e}</div>
                    </div>
                  </div>
                  {warn_html}
                </div>""", unsafe_allow_html=True)
        else:
            st.markdown(
                "<div style='font-size:9px;color:#555;'>No HAZARD nodes have target values set "
                "— set targetValue on HAZARD nodes to see divergence analysis.</div>",
                unsafe_allow_html=True)

        # ── All gate forward values ───────────────────────────────────────
        if result.get("gate_results"):
            with st.expander("All gate forward-computed values (debug)"):
                for gname, prob in result["gate_results"].items():
                    st.markdown(
                        f"<code style='font-size:10px;color:#4fc3f7;'>{esc(gname)}</code>"
                        f"<span style='font-size:10px;color:#ccc;font-family:monospace;'>"
                        f"  &rarr;  {prob:.6e}</span>",
                        unsafe_allow_html=True)


# ── Gist helpers ──────────────────────────────────────────────────────────
def gh(token): return {"Authorization":f"token {token}","Accept":"application/vnd.github+json"}

def get_gist(token, gid):
    try:
        r = requests.get(f"https://api.github.com/gists/{gid}", headers=gh(token), timeout=10)
        return r.json() if r.status_code == 200 else None
    except: return None

def list_gist_files(token, gid):
    g = get_gist(token, gid)
    return sorted(g.get("files",{}).keys()) if g else []

def load_gist_file(token, gid, fname):
    g = get_gist(token, gid)
    if not g: return []
    try:
        raw = json.loads(g.get("files",{}).get(fname,{}).get("content","[]"))
    except:
        return []
    for n in raw:
        n.setdefault("nodeId",    n.get("id",""))
        n.setdefault("ftLabel",   "")
        n.setdefault("fixedValue", None)
        n.setdefault("targetValue", None)
        n.setdefault("calculatedValue", None)
        n.setdefault("parentIds", [])
        n.setdefault("gate", "OR")
        n.setdefault("type", "IF")
        n.setdefault("_pos", None)
    return raw

def save_gist_file(token, gid, fname, data):
    try:
        r = requests.patch(f"https://api.github.com/gists/{gid}", headers=gh(token),
                           json={"files":{fname:{"content":json.dumps(data,indent=2)}}}, timeout=10)
        return r.status_code == 200
    except: return False

def inject_positions(nodes, positions):
    if not positions: return nodes
    updated = []
    for n in nodes:
        n = dict(n)
        if n["id"] in positions:
            n["_pos"] = positions[n["id"]]
        updated.append(n)
    return updated

def extract_positions(nodes):
    return {n["id"]: n["_pos"] for n in nodes if n.get("_pos")}

def del_gist_file(token, gid, fname):
    try:
        requests.patch(f"https://api.github.com/gists/{gid}", headers=gh(token),
                       json={"files":{fname:None}}, timeout=10)
    except: pass

def recalculate(nodes):
    """
    Top-down reverse distribution — fixed version.

    Bugs fixed vs original:
    ─────────────────────────────────────────────────────────────────────────
    FIX 1 — Circular reference detection
        A DFS with white/gray/black coloring runs BEFORE the Kahn queue.
        If a cycle is found, affected nodes are flagged with
        calculatedValue = None and a warning is attached so the UI can
        surface it rather than silently looping or dropping nodes.

    FIX 2 — Negative OR remainder (silent corruption)
        When sum(fixed_children) > parent_val the old code clamped to 0
        and wrote 0.0 to all unfixed children with no warning.
        Now: remainder < 0 is detected, a per-parent warning is recorded,
        and unfixed children are left as None so the UI shows "–" rather
        than a misleading zero.

    FIX 3 — Shared event multi-parent convergence (iterative resolution)
        The old single-pass MAX heuristic was wrong for nodes that appear
        under multiple parents in different hazard branches.  When a shared
        node's parents disagree on the required value the single pass picks
        an arbitrary order and applies MAX on whatever arrived first.
        New approach: after the initial top-down pass, run up to MAX_ITER
        bottom-up / top-down convergence sweeps until every shared node's
        calculatedValue stabilises within CONVERGE_TOL.  This is equivalent
        to the standard FTA iterative importance algorithm.

    FIX 4 — parents_resolved premature enqueue
        The old counter incremented every time ANY parent was processed,
        even if that parent's calculatedValue was still None (because its
        own parent hadn't resolved yet).  A child could therefore enter the
        queue before all valid contributions had arrived.
        Now: a child is only enqueued once ALL parents with a non-None
        calculatedValue have contributed — resolved parents with None are
        not counted.

    Logic unchanged from original:
    ─────────────────────────────────────────────────────────────────────────
    - Pin propagation: same-nodeId group syncs to max(fixedValue)
    - OR gate:  remainder = parent - sum(fixed); each unfixed = remainder/n
    - AND gate: remainder = parent / product(fixed);
                each unfixed = remainder^(1/n)
    - HAZARD seed: calculatedValue = targetValue (or 1e-7 fallback)

    Returns:
        (updated_nodes, warnings)
        updated_nodes — list of node dicts with calculatedValue set
        warnings      — list of human-readable warning strings (may be empty)
    """
    import math
    from collections import defaultdict, deque

    if not nodes:
        return nodes, []

    warnings = []
    updated  = [dict(n) for n in nodes]
    by_id    = {n["id"]: n for n in updated}

    # ── Build children map once ────────────────────────────────────────────
    children_of = {n["id"]: [] for n in updated}
    parent_ids_of = {}
    for n in updated:
        pids = [p for p in (n.get("parentIds") or []) if p in by_id]
        parent_ids_of[n["id"]] = pids
        for pid in pids:
            children_of[pid].append(n["id"])

    # ══════════════════════════════════════════════════════════════════════
    # FIX 1 — Cycle detection (DFS, white/gray/black)
    # ══════════════════════════════════════════════════════════════════════
    WHITE, GRAY, BLACK = 0, 1, 2
    color    = {n["id"]: WHITE for n in updated}
    in_cycle = set()

    def dfs_cycle(nid):
        color[nid] = GRAY
        for cid in children_of.get(nid, []):
            if color[cid] == GRAY:
                # Back-edge found → both endpoints are in a cycle
                in_cycle.add(nid)
                in_cycle.add(cid)
            elif color[cid] == WHITE:
                dfs_cycle(cid)
        color[nid] = BLACK

    for n in updated:
        if color[n["id"]] == WHITE:
            dfs_cycle(n["id"])

    if in_cycle:
        cycle_names = [by_id[nid]["name"] for nid in in_cycle if nid in by_id]
        warnings.append(
            f"CYCLE DETECTED — {len(in_cycle)} node(s) form a loop and will be "
            f"skipped: {', '.join(cycle_names[:8])}"
            + (" …" if len(cycle_names) > 8 else "")
        )

    # ── Pin propagation: sync fixedValue across same-nodeId instances ──────
    nid_groups = defaultdict(list)
    for n in updated:
        nid = (n.get("nodeId") or "").strip()
        if nid:
            nid_groups[nid].append(n)
    for nid, grp in nid_groups.items():
        if len(grp) < 2:
            continue
        pinned_vals = [n["fixedValue"] for n in grp if n.get("fixedValue") is not None]
        if not pinned_vals:
            continue
        group_pin = max(pinned_vals)
        for n in grp:
            n["fixedValue"]      = group_pin
            n["calculatedValue"] = group_pin

    # ── Seed initial values ────────────────────────────────────────────────
    for n in updated:
        if n["id"] in in_cycle:
            n["calculatedValue"] = None          # cycles are left unresolved
        elif n["type"] == "HAZARD":
            n["calculatedValue"] = n.get("targetValue") or 1e-7
        elif n.get("fixedValue") is not None:
            n["calculatedValue"] = n["fixedValue"]
        else:
            n["calculatedValue"] = None

    # ══════════════════════════════════════════════════════════════════════
    # Helper: one top-down distribution pass
    # Returns the set of node IDs that were written/updated this pass.
    # ══════════════════════════════════════════════════════════════════════
    def topdown_pass(by_id, children_of, parent_ids_of, in_cycle, pass_warnings):
        """
        Kahn-style BFS top-down pass.

        FIX 4: A child is only enqueued when the number of RESOLVED parents
        (those whose calculatedValue is not None and are not in a cycle)
        equals the child's total valid-parent count.  This prevents premature
        enqueue when a parent is queued but not yet resolved.
        """
        # Count valid (non-cycle) parents per node
        valid_parent_count = {}
        for n in by_id.values():
            nid = n["id"]
            vp  = [p for p in parent_ids_of.get(nid, [])
                   if p not in in_cycle and by_id.get(p, {}).get("calculatedValue") is not None
                   or by_id.get(p, {}).get("type") == "HAZARD"]
            # Recount: parents that will eventually produce a value
            valid_parent_count[nid] = len([
                p for p in parent_ids_of.get(nid, []) if p not in in_cycle
            ])

        resolved         = set()
        parents_done     = {n["id"]: 0 for n in by_id.values()}
        queue            = deque()
        written          = set()

        # Seed: HAZARD nodes and pinned nodes start the wave
        for n in by_id.values():
            if n["id"] in in_cycle:
                continue
            if n["type"] == "HAZARD" or n.get("fixedValue") is not None:
                resolved.add(n["id"])
                queue.append(n["id"])
            elif valid_parent_count.get(n["id"], 0) == 0:
                # Root non-hazard (orphan) — already resolved as None or fixed
                resolved.add(n["id"])
                queue.append(n["id"])

        while queue:
            pid    = queue.popleft()
            parent = by_id[pid]
            pval   = parent.get("calculatedValue")

            child_ids = [c for c in children_of.get(pid, []) if c not in in_cycle]
            if not child_ids:
                continue
            if pval is None:
                # Parent resolved but has no value — still tick children's counter
                for cid in child_ids:
                    parents_done[cid] += 1
                    if (parents_done[cid] >= valid_parent_count.get(cid, 0)
                            and cid not in resolved):
                        resolved.add(cid)
                        queue.append(cid)
                continue

            fixed_ids   = [c for c in child_ids if by_id[c].get("fixedValue") is not None]
            unfixed_ids = [c for c in child_ids if by_id[c].get("fixedValue") is None]
            n_unfixed   = len(unfixed_ids)

            if parent["gate"] == "OR":
                fixed_sum = sum(by_id[c]["fixedValue"] for c in fixed_ids)
                remainder = pval - fixed_sum

                # ── FIX 2: Negative OR remainder ──────────────────────────
                if remainder < 0:
                    pass_warnings.append(
                        f"⚠ OR gate '{parent['name']}': fixed children sum "
                        f"({fixed_sum:.3e}) exceeds parent target "
                        f"({pval:.3e}).  Unfixed children left unresolved. "
                        f"Reduce fixed values or raise the parent target."
                    )
                    child_val = None
                elif n_unfixed > 0:
                    child_val = remainder / n_unfixed
                else:
                    child_val = None

            else:  # AND
                if fixed_ids:
                    fixed_product = 1.0
                    for c in fixed_ids:
                        fv = by_id[c]["fixedValue"]
                        if fv and fv > 0:
                            fixed_product *= fv
                    if fixed_product <= 0:
                        child_val = None
                    else:
                        remainder = pval / fixed_product
                        child_val = (remainder ** (1.0 / n_unfixed)
                                     if n_unfixed > 0 else None)
                else:
                    child_val = (pval ** (1.0 / n_unfixed)
                                 if n_unfixed > 0 else None)

            for cid in unfixed_ids:
                child    = by_id[cid]
                existing = child.get("calculatedValue")
                if child_val is None:
                    pass  # leave as-is (None or previous value)
                elif existing is None:
                    child["calculatedValue"] = child_val
                    written.add(cid)
                else:
                    # Conservative MAX for shared events (first-pass heuristic;
                    # iterative convergence in outer loop corrects this)
                    new_val = max(existing, child_val)
                    if new_val != existing:
                        written.add(cid)
                    child["calculatedValue"] = new_val

            # FIX 4: increment counter only after this parent contributes
            for cid in child_ids:
                parents_done[cid] += 1
                vp_count = valid_parent_count.get(cid, 0)
                if parents_done[cid] >= vp_count and cid not in resolved:
                    resolved.add(cid)
                    queue.append(cid)

        return written

    # ══════════════════════════════════════════════════════════════════════
    # FIX 3 — Iterative convergence for shared events
    # ══════════════════════════════════════════════════════════════════════
    MAX_ITER     = 8
    CONVERGE_TOL = 1e-9   # relative tolerance

    pass_warnings = []

    for iteration in range(MAX_ITER):
        # Snapshot values before this pass
        before = {n["id"]: n.get("calculatedValue") for n in updated}

        topdown_pass(by_id, children_of, parent_ids_of, in_cycle, pass_warnings)

        # Check convergence: all shared nodes stable?
        max_rel_change = 0.0
        for n in updated:
            if n["id"] in in_cycle:
                continue
            old = before.get(n["id"])
            new = n.get("calculatedValue")
            if old is None or new is None:
                continue
            if old == 0.0:
                if new != 0.0:
                    max_rel_change = max(max_rel_change, 1.0)
            else:
                max_rel_change = max(max_rel_change, abs(new - old) / abs(old))

        if max_rel_change <= CONVERGE_TOL:
            break   # converged
    else:
        # Didn't fully converge — warn but keep the best approximation
        warnings.append(
            f"Iterative convergence did not fully settle after {MAX_ITER} passes "
            f"(max relative change: {max_rel_change:.2e}). "
            f"Results are approximate — check for high-probability shared events."
        )

    # Surface any per-pass warnings (deduplicated)
    seen_warn = set()
    for w in pass_warnings:
        if w not in seen_warn:
            warnings.append(w)
            seen_warn.add(w)

    # Final sanity: clamp any value that slipped above 1.0 (floating-point drift)
    for n in updated:
        v = n.get("calculatedValue")
        if v is not None and not math.isnan(v) and not math.isinf(v):
            n["calculatedValue"] = min(v, 1.0)

    return updated, warnings

def fmt(v):
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))): return "-"
    return f"{v:.3e}"

def now_str(): return datetime.now().strftime("%Y-%m-%d_%H-%M")
def is_snap(n): return n.startswith("snapshot_")
def is_named(n): return not is_snap(n)

# ── Export: JSON ──────────────────────────────────────────────────────────
def export_json(nodes):
    clean = []
    for n in nodes:
        nc = dict(n)
        nc.setdefault("nodeId",   nc.get("id",""))
        nc.setdefault("ftLabel",  "")
        nc.setdefault("fixedValue", None)
        nc.setdefault("targetValue", None)
        nc.setdefault("calculatedValue", None)
        clean.append(nc)
    return json.dumps(clean, indent=2).encode("utf-8")

# ── Export: Cypher (Neo4j) ────────────────────────────────────────────────
def export_cypher(nodes):
    by_id = {n["id"]: n for n in nodes}
    n_pinned = sum(1 for n in nodes if n.get("fixedValue") is not None)
    n_shared  = sum(1 for n in nodes if len(n.get("parentIds") or []) > 1)
    lines = [
        "// ── FTA Fault Tree — Cypher Export ──────────────────────────",
        f"// Generated:    {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"// Total nodes:  {len(nodes)}",
        f"// Pinned nodes: {n_pinned}",
        f"// Shared nodes: {n_shared}",
        "//",
        "// MATCH (n:FTANode) DETACH DELETE n;",
        "",
        "// STEP 2: Create all nodes",
    ]
    for n in nodes:
        val       = n.get("calculatedValue")
        fv        = n.get("fixedValue")
        val_s     = fmt(val) if val is not None else "null"
        fv_s      = fmt(fv)  if fv  is not None else "null"
        name      = n["name"].replace("'", "\\'")
        node_id   = (n.get("nodeId") or n["id"]).replace("'", "\\'")
        ft_lbl    = (n.get("ftLabel") or "").replace("'", "\\'")
        is_shared = len(n.get("parentIds") or []) > 1
        is_pinned = fv is not None
        is_group  = n["type"] == "GROUP"
        lines.append(
            f"CREATE (:FTANode {{"
            f"id:'{n['id']}', nodeId:'{node_id}', ftLabel:'{ft_lbl}', name:'{name}', "
            f"type:'{n['type']}', gate:'{n['gate']}', "
            f"calculatedValue:{val_s if val is not None else 'null'}, "
            f"fixedValue:{fv_s if fv is not None else 'null'}, "
            f"valueStr:'{val_s}', "
            f"shared:{str(is_shared).lower()}, "
            f"pinned:{str(is_pinned).lower()}, "
            f"isGroup:{str(is_group).lower()}"
            f"}});"
        )
    lines += ["", "// STEP 3: Create relationships"]
    for n in nodes:
        for pid in (n.get("parentIds") or []):
            if pid in by_id:
                lines.append(
                    f"MATCH (c:FTANode {{id:'{n['id']}'}}), (p:FTANode {{id:'{pid}'}}) "
                    f"CREATE (c)-[:FEEDS_INTO {{gate:'{by_id[pid]['gate']}', "
                    f"childPinned:{str(n.get('fixedValue') is not None).lower()}}}]->(p);"
                )
    return "\n".join(lines).encode("utf-8")

# ── Export: Excel ─────────────────────────────────────────────────────────
def sanitize_xl(val):
    if val is None: return "-"
    if isinstance(val, (int, float)): return val
    s = str(val)
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        s = ILLEGAL_CHARACTERS_RE.sub("", s)
    except ImportError: pass
    return s or "-"

def export_excel(nodes):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError: return None

    by_id    = {n["id"]: n for n in nodes}
    wb       = openpyxl.Workbook()
    fills    = {lvl: PatternFill("solid", fgColor=c) for lvl, c in
                [("HAZARD","FFFF4D4D"),("SF","FFFF8C42"),("FF","FFF5C518"),
                 ("IF","FF4CAF7D"),("GROUP","FF7E57C2")]}
    hdr_fill = PatternFill("solid", fgColor="FF0F3460")
    pin_fill = PatternFill("solid", fgColor="FF3A0A14")

    def hdr_font(c): return Font(bold=True, color="FFFFFFFF", name="Courier New")
    def row_font(lvl):
        dark_text = lvl in ("FF",)
        return Font(name="Courier New", size=10, color="FF111111" if dark_text else "FFFFFFFF")
    def ctr(): return Alignment(horizontal="center", vertical="center")
    def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws = wb.active; ws.title = "FTA Nodes"
    hdrs = ["Level", "Type", "Node Name", "Node ID", "FT Label",
            "Gate", "Calc. Value", "Fixed Value (📌)", "Shared",
            "Parent Nodes", "Child Nodes"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(1, ci, h); c.font = hdr_font(None); c.fill = hdr_fill; c.alignment = ctr()
    row = 2
    for lvl in DISPLAY_ORDER:
        for n in [x for x in nodes if x["type"] == lvl]:
            pnames   = " | ".join(by_id[p]["name"] for p in (n.get("parentIds") or []) if p in by_id)
            cnames   = " | ".join(x["name"] for x in nodes if n["id"] in (x.get("parentIds") or []))
            fv       = n.get("fixedValue")
            fv_str   = fmt(fv) if fv is not None else ""
            is_pinned = fv is not None
            ft_lbl   = n.get("ftLabel", "")
            vals = [
                DISPLAY_ORDER.index(lvl) + 1,
                sanitize_xl(lvl),
                sanitize_xl(n["name"]),
                sanitize_xl(n.get("nodeId", n["id"])),
                sanitize_xl(ft_lbl),
                sanitize_xl(n["gate"]),
                n.get("calculatedValue"),
                sanitize_xl(fv_str),
                "YES" if len(n.get("parentIds") or []) > 1 else "NO",
                sanitize_xl(pnames or "-"),
                sanitize_xl(cnames or "-"),
            ]
            node_fill = pin_fill if is_pinned else fills.get(lvl, hdr_fill)
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row, ci, sanitize_xl(v) if isinstance(v, str) else v)
                cell.font      = row_font(lvl)
                cell.fill      = node_fill
                cell.alignment = lft() if ci in (3, 10, 11) else ctr()
            if is_pinned:
                ws.cell(row, 8).font = Font(name="Courier New", size=10,
                                            color="FFFF4D4D", bold=True)
            row += 1
    for ci, w in enumerate([8, 10, 30, 12, 10, 8, 16, 16, 8, 30, 30], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("Hierarchy"); ws2.sheet_view.showGridLines = False
    ws2.cell(1, 1, "FTA HIERARCHY - TOP TO BOTTOM").font = Font(
        bold=True, size=14, name="Courier New", color="FFE94560")
    ws2.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        size=9, name="Courier New", color="FF888888")
    row = 4
    def write_hier(nid, depth, seen):
        nonlocal row
        if nid in seen: return
        seen.add(nid)
        n = by_id.get(nid)
        if not n: return
        label  = f"{'    ' * depth}{'  -> ' if depth else ''}{n['name']}"
        tags   = []
        nid_tag = n.get("nodeId", "")
        ft_tag  = n.get("ftLabel", "")
        if nid_tag:                                tags.append(f"[{nid_tag}]")
        if ft_tag:                                 tags.append(f"[{ft_tag}]")
        if len(n.get("parentIds") or []) > 1:      tags.append("[SHARED]")
        if n.get("fixedValue") is not None:        tags.append(f"[FIXED={fmt(n['fixedValue'])}]")
        if n["type"] == "GROUP":                   tags.append("[GROUP]")
        tag_str = " ".join(tags)
        fhex = {"HAZARD":"FFFF4D4D","SF":"FFFF8C42","FF":"FFF5C518",
                "IF":"FF4CAF7D","GROUP":"FF7E57C2"}.get(n["type"], "FF1A3A6B")
        dark_text = n["type"] in ("FF",)
        txt_color = "FF111111" if dark_text else "FFFFFFFF"
        c1 = ws2.cell(row, 1, sanitize_xl(label))
        c2 = ws2.cell(row, 2, sanitize_xl(f"{n['type']}[{n['gate']}]"))
        c3 = ws2.cell(row, 3, sanitize_xl(fmt(n.get("calculatedValue"))))
        c4 = ws2.cell(row, 4, sanitize_xl(tag_str))
        for c in [c1, c2, c3, c4]:
            c.font = Font(name="Courier New", size=10, bold=(depth == 0), color=txt_color)
            c.fill = PatternFill("solid", fgColor=fhex)
        c3.alignment = Alignment(horizontal="right", vertical="center")
        if n.get("fixedValue") is not None:
            c3.font = Font(name="Courier New", size=10, bold=True, color="FFFF4D4D")
        row += 1
        for child in [x for x in nodes if nid in (x.get("parentIds") or [])]:
            write_hier(child["id"], depth + 1, seen)

    for h in [n for n in nodes if n["type"] == "HAZARD"]:
        write_hier(h["id"], 0, set())
        row += 1
    for ci, w in enumerate([50, 16, 16, 24], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ── Tree HTML builder ─────────────────────────────────────────────────────
def build_html_tree(nodes, filter_hazard_id=None, tree_state=None):
    if not nodes: return ""
    import json as _json

    by_id = {n["id"]: n for n in nodes}
    if filter_hazard_id:
        visible = set()
        q = [filter_hazard_id]
        while q:
            cur = q.pop()
            if cur in visible: continue
            visible.add(cur)
            for child in [n for n in nodes if cur in (n.get("parentIds") or [])]:
                q.append(child["id"])
        for n in nodes:
            if n["type"] != "HAZARD" and not n.get("parentIds"):
                _fq = [c["id"] for c in nodes if n["id"] in (c.get("parentIds") or [])]
                _seen = set()
                _found = False
                while _fq and not _found:
                    _cid = _fq.pop()
                    if _cid in _seen: continue
                    _seen.add(_cid)
                    if _cid in visible:
                        _found = True
                        break
                    _fq += [c["id"] for c in nodes if _cid in (c.get("parentIds") or [])]
                if _found:
                    visible.add(n["id"])
        show_nodes = [n for n in nodes if n["id"] in visible]
    else:
        show_nodes = nodes
    if not show_nodes: return ""

    shown_ids  = {n["id"] for n in show_nodes}
    shared_ids = {n["id"] for n in show_nodes
                  if len([p for p in (n.get("parentIds") or []) if p in shown_ids]) > 1}

    from collections import defaultdict as _dd
    nid_groups = _dd(list)
    for n in show_nodes:
        nid = (n.get("nodeId") or "").strip()
        if nid:
            nid_groups[nid].append(n["id"])
    duplicate_ids = {iid for group in nid_groups.values() if len(group) > 1 for iid in group}

    ts        = tree_state or {}
    init_pos  = dict(ts.get("positions", {}))
    focus_id  = ts.get("focus_id")

    for n in show_nodes:
        if n.get("_pos") and n["id"] not in init_pos:
            init_pos[n["id"]] = n["_pos"]

    LEVEL_ROW   = {"HAZARD": 0, "SF": 1, "FF": 2, "GROUP": 2.5, "IF": 3}
    LEVEL_COLOR = {0: "#ff4d4d", 1: "#ff8c42", 2: "#f5c518", 2.5: "#7e57c2", 3: "#4caf7d"}
    LEVEL_LABEL = {0: "HAZARD", 1: "SF", 2: "FF", 2.5: "GROUP", 3: "IF"}

    _depth_map = {}
    def _get_depth(nid, _seen=None):
        if nid in _depth_map: return _depth_map[nid]
        if _seen is None: _seen = set()
        if nid in _seen: return 0
        _seen = _seen | {nid}
        node = next((n for n in show_nodes if n["id"] == nid), None)
        if node is None: return 0
        pids = [p for p in (node.get("parentIds") or []) if p in shown_ids]
        if not pids:
            _depth_map[nid] = 0
        else:
            _depth_map[nid] = max(_get_depth(p, _seen) for p in pids) + 1
        return _depth_map[nid]
    for n in show_nodes:
        _get_depth(n["id"])

    nodes_js = _json.dumps([{
        "id":          n["id"],
        "name":        n["name"],
        "type":        n["type"],
        "gate":        n["gate"],
        "value":       fmt(n.get("calculatedValue")),
        "nodeId":      n.get("nodeId", n["id"]),
        "ftLabel":     n.get("ftLabel", ""),
        "shared":      n["id"] in shared_ids,
        "isDuplicate": n["id"] in duplicate_ids,
        "isPinned":    n.get("fixedValue") is not None,
        "fixedVal":    fmt(n.get("fixedValue")) if n.get("fixedValue") is not None else None,
        "isGroup":     n["type"] == "GROUP",
        "isRoot":      n["type"] != "HAZARD" and not [p for p in (n.get("parentIds") or []) if p in shown_ids],
        "color":       LEVEL_COLORS.get(n["type"], "#7e57c2"),
        "tcolor":      LEVEL_TEXT.get(n["type"], "#fff"),
        "row":         _depth_map.get(n["id"], LEVEL_ROW.get(n["type"], 2)),
        "parents":     [p for p in (n.get("parentIds") or []) if p in shown_ids],
        "pnames":      [by_id[p]["name"] for p in (n.get("parentIds") or []) if p in shown_ids],
        "children":    [c["id"] for c in show_nodes if n["id"] in (c.get("parentIds") or [])],
        "cnames":      [c["name"] for c in show_nodes if n["id"] in (c.get("parentIds") or [])],
    } for n in show_nodes])

    links_js = _json.dumps([
        {"sid": pid, "tid": n["id"],
         "andGate": by_id[pid]["gate"] == "AND" if pid in by_id else False,
         "shared":  n["id"] in shared_ids}
        for n in show_nodes
        for pid in (n.get("parentIds") or [])
        if pid in shown_ids
    ])

    init_pos_js    = _json.dumps(init_pos)
    focus_js       = f'"{focus_id}"' if focus_id else "null"
    level_label_js = _json.dumps({str(k): v for k, v in LEVEL_LABEL.items()})
    level_color_js = _json.dumps({str(k): v for k, v in LEVEL_COLOR.items()})

    # ── D3 HTML tree ─────────────────────────────────────────────────────
    html = """<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#0a0a0a;font-family:'JetBrains Mono','Fira Code',monospace;
     color:#e0e0e0;overflow:hidden;height:100vh;display:flex;flex-direction:column;}
#toolbar{display:flex;align-items:center;gap:6px;padding:6px 12px;
  background:#111;border-bottom:2px solid #1a1a1a;flex-shrink:0;user-select:none;height:44px;}
.sep{width:1px;height:22px;background:#2a2a2a;flex-shrink:0;}
.btn{background:#1c1c1c;border:1.5px solid #2e2e2e;color:#bbb;border-radius:6px;
  padding:0 11px;cursor:pointer;font-family:inherit;font-size:11px;font-weight:700;
  letter-spacing:.4px;transition:.12s;white-space:nowrap;flex-shrink:0;
  height:28px;display:inline-flex;align-items:center;gap:4px;}
.btn:hover{background:#282828;color:#fff;border-color:#555;}
.btn.on{background:#e94560;border-color:#e94560;color:#fff;}
#zlbl{color:#666;font-size:11px;min-width:40px;text-align:center;font-weight:700;}
#fst{font-size:10px;color:#f5c518;padding:2px 8px;border-radius:5px;
  background:#1a1300;border:1px solid #f5c51844;white-space:nowrap;display:none;}
#fst.show{display:inline-block;}
#swrap{display:flex;align-items:center;gap:4px;margin-left:auto;max-width:260px;}
#sbox{background:#1c1c1c;border:1.5px solid #2e2e2e;color:#ccc;border-radius:6px;
  padding:0 9px;font-family:inherit;font-size:11px;outline:none;width:155px;height:28px;}
#sbox:focus{border-color:#e94560;color:#fff;}
#sbox::placeholder{color:#444;}
#si{color:#666;font-size:10px;min-width:50px;}
/* Main layout: canvas + right panel */
#main{flex:1;display:flex;overflow:hidden;}
#wrap{flex:1;position:relative;overflow:hidden;}
svg{position:absolute;inset:0;width:100%;height:100%;}
#lanes{position:absolute;inset:0;pointer-events:none;z-index:2;overflow:hidden;}
.lb{position:absolute;left:0;right:0;border-top:1px solid rgba(255,255,255,.04);
  display:flex;align-items:flex-start;padding-top:5px;}
.lt{margin-left:10px;font-size:8px;font-weight:700;letter-spacing:2.5px;opacity:.4;}
.ls{position:absolute;left:0;top:0;width:3px;height:100%;opacity:.6;}
/* Right node panel */
#rp{width:0;min-width:0;background:#0d0d0d;border-left:2px solid #1a1a1a;
  display:flex;flex-direction:column;transition:width .2s ease,min-width .2s ease;
  overflow:hidden;flex-shrink:0;}
#rp.open{width:290px;min-width:290px;}
#rp-inner{width:290px;padding:10px 12px;overflow-y:auto;height:100%;}
.rp-hdr{font-size:8px;font-weight:700;letter-spacing:2px;color:#555;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid #1a1a1a;}
.rp-row{margin-bottom:8px;}
.rp-lbl{font-size:7px;color:#444;letter-spacing:1.5px;font-weight:700;margin-bottom:2px;}
.rp-val{font-size:11px;font-weight:700;font-family:monospace;word-break:break-all;}
.rp-input{background:#1a1a1a;border:1.5px solid #2e2e2e;color:#ddd;border-radius:5px;
  padding:4px 8px;font-family:inherit;font-size:11px;outline:none;width:100%;margin-top:2px;}
.rp-input:focus{border-color:#e94560;}
.rp-sel{background:#1a1a1a;border:1.5px solid #2e2e2e;color:#ddd;border-radius:5px;
  padding:4px 8px;font-family:inherit;font-size:11px;outline:none;width:100%;margin-top:2px;}
.rp-btn{background:#1c1c1c;border:1.5px solid #2e2e2e;color:#bbb;border-radius:5px;
  padding:5px 10px;cursor:pointer;font-family:inherit;font-size:10px;font-weight:700;
  width:100%;margin-top:4px;transition:.12s;}
.rp-btn:hover{background:#282828;color:#fff;border-color:#555;}
.rp-btn.primary{background:#e9456033;border-color:#e94560;color:#e94560;}
.rp-btn.primary:hover{background:#e9456055;}
.rp-btn.save{background:#4caf7d33;border-color:#4caf7d;color:#4caf7d;}
.rp-btn.save:hover{background:#4caf7d55;}
.rp-chip{display:inline-block;padding:1px 6px;border-radius:10px;font-size:9px;
  font-family:monospace;font-weight:700;margin:1px 2px;cursor:pointer;}
#rp-close{position:absolute;top:8px;right:8px;background:none;border:none;color:#444;
  font-size:15px;cursor:pointer;padding:2px 5px;}
#rp-close:hover{color:#fff;}
/* Multi-select bar */
#msp{display:none;position:absolute;bottom:0;left:0;right:0;
  background:rgba(8,8,8,.95);border-top:2px solid #4fc3f7;padding:8px 14px 10px;
  z-index:21;backdrop-filter:blur(14px);}
</style>
</head><body>
<div id="toolbar">
  <button class="btn" onclick="zBy(.22)">&#65291;</button>
  <button class="btn" onclick="zBy(-.22)">&#65293;</button>
  <span id="zlbl">85%</span>
  <div class="sep"></div>
  <button class="btn" id="blay" onclick="doColumnLayout(true)">&#8862; Reset</button>
  <button class="btn" onclick="doFit()">&#8865; Fit</button>
  <div class="sep"></div>
  <button class="btn" id="bfrc" onclick="toggleForce()">&#9889; Force</button>
  <span id="fst"></span>
  <div class="sep"></div>
  <button class="btn" onclick="clearHL()">&#10005; Clear</button>
  <div id="swrap">
    <input id="sbox" placeholder="Search&#8230;" oninput="doSearch(this.value)"/>
    <button class="btn" onclick="sPrev()">&#9650;</button>
    <button class="btn" onclick="sNext()">&#9660;</button>
    <span id="si"></span>
  </div>
</div>
<div id="main">
<div id="wrap">
  <div id="lanes"></div>
  <svg id="sv">
    <defs>
      <!-- FTA standard gate markers: arrows point from child UP to parent -->
      <marker id="arr" markerWidth="9" markerHeight="9" refX="8" refY="4.5" orient="auto">
        <path d="M0,0 L9,4.5 L0,9 Z" fill="#3a3a3a"/>
      </marker>
      <marker id="arr-or" markerWidth="9" markerHeight="9" refX="8" refY="4.5" orient="auto">
        <path d="M0,0 L9,4.5 L0,9 Z" fill="#4fc3f7"/>
      </marker>
      <marker id="arr-and" markerWidth="9" markerHeight="9" refX="8" refY="4.5" orient="auto">
        <path d="M0,0 L9,4.5 L0,9 Z" fill="#ffb74d"/>
      </marker>
      <marker id="arr-sh" markerWidth="9" markerHeight="9" refX="8" refY="4.5" orient="auto">
        <path d="M0,0 L9,4.5 L0,9 Z" fill="#f5c518"/>
      </marker>
    </defs>
    <g id="zg"><g id="lg"></g><g id="gg"></g><g id="ng"></g></g>
  </svg>
  <div id="msp">
    <button onclick="clearMultiSel()" style="position:absolute;top:6px;right:10px;
      background:none;border:none;color:#555;font-size:15px;cursor:pointer;">&#10005;</button>
    <div style="font-size:8px;color:#4fc3f7;letter-spacing:3px;margin-bottom:4px;font-weight:700;">
      MULTI-SELECT — <span id="msc">0</span> NODES
    </div>
    <div id="mslist" style="display:flex;flex-wrap:wrap;gap:3px;max-height:44px;overflow-y:auto;"></div>
  </div>
</div>
<!-- Right Panel: node inspector + editor -->
<div id="rp">
  <div id="rp-inner">
    <button id="rp-close" onclick="closeRP()">&#10005;</button>
    <div style="font-size:8px;color:#555;letter-spacing:3px;margin-bottom:6px;margin-right:20px;">NODE INSPECTOR</div>
    <!-- Node header -->
    <div id="rp-name" style="font-size:13px;font-weight:700;margin-bottom:2px;word-break:break-word;"></div>
    <div id="rp-badge" style="margin-bottom:10px;"></div>
    <!-- Stats grid -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:10px;">
      <div style="background:#111;border-radius:5px;padding:6px;">
        <div class="rp-lbl">TYPE</div>
        <div class="rp-val" id="rp-type"></div>
      </div>
      <div style="background:#111;border-radius:5px;padding:6px;">
        <div class="rp-lbl">GATE</div>
        <div class="rp-val" id="rp-gate-disp"></div>
      </div>
      <div style="background:#111;border-radius:5px;padding:6px;">
        <div class="rp-lbl">CALC VALUE</div>
        <div class="rp-val" id="rp-value"></div>
      </div>
      <div style="background:#111;border-radius:5px;padding:6px;">
        <div class="rp-lbl">NODE ID</div>
        <div class="rp-val" id="rp-nid" style="font-size:10px;"></div>
      </div>
    </div>
    <!-- Parents / Children -->
    <div style="margin-bottom:8px;">
      <div class="rp-lbl">&#8593; PARENTS</div>
      <div id="rp-parents" style="font-size:10px;color:#888;line-height:1.6;margin-top:2px;"></div>
    </div>
    <div style="margin-bottom:10px;">
      <div class="rp-lbl">&#8595; CHILDREN (<span id="rp-ifcount" style="color:#4caf7d"></span>)</div>
      <div id="rp-children" style="font-size:10px;color:#888;line-height:1.6;margin-top:2px;"></div>
    </div>
    <hr style="border-color:#1a1a1a;margin:8px 0;">
    <!-- Edit section -->
    <div style="font-size:8px;color:#555;letter-spacing:2px;margin-bottom:6px;font-weight:700;">EDIT NODE</div>
    <div class="rp-row">
      <div class="rp-lbl">NAME</div>
      <input class="rp-input" id="rp-edit-name" type="text" placeholder="Node name"/>
    </div>
    <div class="rp-row">
      <div class="rp-lbl">GATE TYPE</div>
      <select class="rp-sel" id="rp-edit-gate">
        <option value="OR">OR</option>
        <option value="AND">AND</option>
      </select>
    </div>
    <div class="rp-row">
      <div class="rp-lbl">NODE ID</div>
      <input class="rp-input" id="rp-edit-nodeid" type="text" placeholder="e.g. IF-042"/>
    </div>
    <div class="rp-row">
      <div class="rp-lbl">FT LABEL</div>
      <input class="rp-input" id="rp-edit-ftlabel" type="text" placeholder="e.g. FT42"/>
    </div>
    <div class="rp-row">
      <div class="rp-lbl">FIXED VALUE (PIN &#128204;) — leave blank to unpin</div>
      <input class="rp-input" id="rp-edit-fixed" type="text" placeholder="e.g. 1.5e-5"/>
    </div>
    <div class="rp-row">
      <div class="rp-lbl">TARGET VALUE (HAZARD only)</div>
      <input class="rp-input" id="rp-edit-target" type="text" placeholder="e.g. 1e-7"/>
    </div>
    <button class="rp-btn save" onclick="rpSave()">&#10003; Save Changes</button>
    <div id="rp-save-msg" style="font-size:9px;color:#4caf7d;margin-top:4px;display:none;">Saved!</div>
    <hr style="border-color:#1a1a1a;margin:8px 0;">
    <button class="rp-btn primary" onclick="rpDelete()" id="rp-del-btn">&#128465; Delete Node</button>
  </div>
</div>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.8.5/d3.min.js"></script>
<script>
const RNODES=__NODES__;
const RLINKS=__LINKS__;
const IPOS=__IPOS__;
const FOCUSID=__FOCUS__;
const LLABELS=__LLABELS__;
const LCOLORS=__LCOLORS__;
const GC={OR:"#4fc3f7",AND:"#ffb74d"};
const NW=196,NH=96,HG=50,VG=210;
let selId=null,forceOn=false,forceSub=null;
let collapsed=new Set(),sM=[],sI=0;
let multiSel=new Set();
const uP={};
Object.entries(IPOS).forEach(([id,p])=>uP[id]={x:p.x,y:p.y});
const NM={};
RNODES.forEach(n=>NM[n.id]=n);
const wrap=document.getElementById("wrap");
const svg=d3.select("#sv");
const zg=svg.select("#zg"),lg=svg.select("#lg"),gg=svg.select("#gg"),ng=svg.select("#ng");
const zb=d3.zoom().scaleExtent([.03,6])
  .filter(e=>e.button===2||e.type==="wheel")
  .on("zoom",e=>{zg.attr("transform",e.transform);document.getElementById("zlbl").textContent=Math.round(e.transform.k*100)+"%";updateLanes();});
svg.call(zb).on("contextmenu",e=>e.preventDefault());
svg.on("click",()=>{if(multiSel.size>0)clearMultiSel();});
function getT(){return d3.zoomTransform(svg.node());}

// ── Layer rows ─────────────────────────────────────────────────────────────
// FTA top-down: HAZARD at top (y=0), children below
const TYPE_ROW={HAZARD:0,SF:1,FF:2,GROUP:2,IF:3};
function getNodeRow(n){
  if(n._midRow!=null) return n._midRow;
  return TYPE_ROW[n.type]??2;
}
function layY(n){
  const r=typeof n==="object"?getNodeRow(n):(TYPE_ROW[n]??2);
  return 80+r*VG;
}

let sN=[],sL=[];
const sim=d3.forceSimulation()
  .force("link",d3.forceLink().id(d=>d.id).distance(NW+HG).strength(.2))
  .force("charge",d3.forceManyBody().strength(-1400).distanceMax(800))
  .force("collide",d3.forceCollide(NW*0.72))
  .force("y",d3.forceY(d=>layY(d)).strength(3.5))
  .force("x",d3.forceX(d=>d.bx||600).strength(.2))
  .alphaDecay(.015).velocityDecay(.6)
  .on("tick",tick);
sim.stop();

function getVis(){
  const h=new Set();
  collapsed.forEach(cid=>{const q=[cid];while(q.length){const c=q.shift();(NM[c]?.children||[]).forEach(ch=>{if(!h.has(ch)){h.add(ch);q.push(ch);}});}});
  return RNODES.filter(n=>!h.has(n.id));
}
function subIds(id){
  const s=new Set([id]),q=[id];
  while(q.length){const c=q.shift();(NM[c]?.children||[]).forEach(ch=>{if(!s.has(ch)&&sN.find(n=>n.id===ch)){s.add(ch);q.push(ch);}});}
  return s;
}
function countIFs(id){
  const seen=new Set(),q=[id]; let c=0;
  while(q.length){const x=q.shift();if(seen.has(x))continue;seen.add(x);const n=NM[x];if(!n)continue;if(n.type==="IF")c++;(n.children||[]).forEach(ch=>q.push(ch));}
  return c;
}

function refresh(){
  const vis=getVis(); const ex={};
  sN.forEach(n=>ex[n.id]=n);
  sN=vis.map(n=>Object.assign({...n},{
    x:ex[n.id]?.x??uP[n.id]?.x??null,
    y:ex[n.id]?.y??uP[n.id]?.y??null,
    vx:0,vy:0,fx:null,fy:null,bx:null,_midRow:null
  }));
  const vs=new Set(sN.map(n=>n.id));
  // Links go from PARENT (source=sid) to CHILD (target=tid)
  // Arrow drawn at target (child) end — visual flow: parent→child top-down
  // But FTA convention: child causes parent — so arrow from child to parent
  // We flip: source=tid (child), target=sid (parent) for arrow direction
  sL=RLINKS.filter(l=>vs.has(l.sid)&&vs.has(l.tid)).map(l=>{
    const child=sN.find(n=>n.id===l.tid);   // child node
    const parent=sN.find(n=>n.id===l.sid);  // parent node
    return child&&parent?{source:child,target:parent,andGate:l.andGate,shared:l.shared,parentId:l.sid,childId:l.tid}:null;
  }).filter(Boolean);
  assignMidRows();
  computeRTLayout(false);
  sim.nodes(sN); sim.force("link").links(sL); sim.force("x").x(d=>d.bx||600);
  drawLinks(); drawGateSymbols(); drawNodes(); tick(); updateLanes();
  if(forceOn) sim.alpha(.3).restart();
}

function assignMidRows(){
  const sfAncestors={};
  sN.forEach(n=>{sfAncestors[n.id]=new Set();});
  sN.filter(n=>n.type==="SF").forEach(sf=>{
    const q=[sf.id];const seen=new Set();
    while(q.length){
      const cid=q.shift();if(seen.has(cid))continue;seen.add(cid);
      if(cid!==sf.id){if(!sfAncestors[cid])sfAncestors[cid]=new Set();sfAncestors[cid].add(sf.id);}
      const n=sN.find(x=>x.id===cid);if(n)(n.children||[]).forEach(ch=>q.push(ch));
    }
  });
  sN.forEach(n=>{
    const s=sfAncestors[n.id]||new Set();
    n._midRow=(n.type==="FF"||n.type==="GROUP")&&s.size>1?1.5:null;
  });
}

// ── Links: draw FROM child (bottom) TO parent (top) ───────────────────────
// source = child, target = parent — line goes bottom-to-top with arrow at parent
function drawLinks(){
  const s=lg.selectAll("path.lk").data(sL,d=>d.childId+"->"+d.parentId);
  const a=s.enter().append("path").attr("class","lk").attr("fill","none").merge(s);
  a.attr("stroke",d=>{
    if(d.shared) return "#f5c518bb";
    return d.andGate?"#ffb74d99":"#4fc3f766";
  })
   .attr("stroke-width",d=>d.shared?2:2.5)
   .attr("stroke-dasharray",d=>d.andGate?"8,4":null)
   .attr("marker-end",d=>{
     if(d.shared)  return "url(#arr-sh)";
     return d.andGate?"url(#arr-and)":"url(#arr-or)";
   });
  s.exit().remove();
}

// ── FTA standard gate symbols ──────────────────────────────────────────────
// OR gate: curved shield shape (standard FTA)
// AND gate: flat-top D shape (standard FTA)
// Placed at the parent node's bottom, one per parent showing its gate type
function drawGateSymbols(){
  // Build per-parent gate data (one gate per parent node)
  const parentGates=new Map(); // parentId → {node, gate, x, y}
  sL.forEach(l=>{
    const pid=l.parentId;
    if(!parentGates.has(pid)){
      const pn=sN.find(n=>n.id===pid);
      if(pn) parentGates.set(pid,{n:pn,andGate:l.andGate});
    }
  });
  const gdata=Array.from(parentGates.values());

  const s=gg.selectAll("g.gsym").data(gdata,d=>d.n.id);
  const e=s.enter().append("g").attr("class","gsym");
  // Gate body path (will be set per type)
  e.append("path").attr("class","gshape");
  e.append("text").attr("class","glbl")
    .attr("text-anchor","middle").attr("dominant-baseline","central")
    .attr("font-size","7px").attr("font-weight","700").attr("font-family","monospace");
  const all=e.merge(s);

  // Standard FTA gate sizes
  const GW=22,GH=16;
  // OR gate: curved arc (shield) path at origin, centred
  function orPath(){
    return `M${-GW/2},${-GH/2} Q0,${-GH/2-6} ${GW/2},${-GH/2} Q${GW/2+4},0 ${GW/2},${GH/2} Q0,${GH/2+8} ${-GW/2},${GH/2} Q${-GW/2-4},0 ${-GW/2},${-GH/2} Z`;
  }
  // AND gate: flat top, curved bottom
  function andPath(){
    return `M${-GW/2},${-GH/2} L${GW/2},${-GH/2} Q${GW/2+4},0 ${GW/2},${GH/2} Q0,${GH/2+6} ${-GW/2},${GH/2} Z`;
  }

  all.select("path.gshape")
    .attr("d",d=>d.andGate?andPath():orPath())
    .attr("fill",d=>d.andGate?"#1a0e00":"#001216")
    .attr("stroke",d=>d.andGate?"#ffb74d":"#4fc3f7")
    .attr("stroke-width","1.8");
  all.select("text.glbl")
    .text(d=>d.andGate?"AND":"OR")
    .attr("fill",d=>d.andGate?"#ffb74d":"#4fc3f7");

  s.exit().remove();
}

// ── Nodes ──────────────────────────────────────────────────────────────────
function drawNodes(){
  const s=ng.selectAll("g.nd").data(sN,d=>d.id);
  const e=s.enter().append("g").attr("class","nd").style("cursor","pointer");
  e.append("rect").attr("class","nb").attr("width",NW).attr("height",NH).attr("rx",7);
  // Type label top-left
  e.append("text").attr("class","nt").attr("x",8).attr("y",14)
   .attr("font-size","7px").attr("letter-spacing","1.5px").attr("font-weight","700");
  // Gate label top-right
  e.append("text").attr("class","ngt").attr("x",NW-8).attr("y",14)
   .attr("text-anchor","end").attr("font-size","7px").attr("font-weight","700");
  // ID prefix (colored) + ID number — row 2
  e.append("text").attr("class","nidp").attr("x",8).attr("y",30)
   .attr("font-size","9px").attr("font-weight","700").attr("font-family","monospace");
  e.append("text").attr("class","nidn").attr("x",0).attr("y",30)
   .attr("font-size","9px").attr("font-family","monospace").attr("fill","#aaa");
  // Node name — row 3, bold
  e.append("text").attr("class","nn").attr("x",NW/2).attr("y",50)
   .attr("text-anchor","middle").attr("font-size","11px").attr("font-weight","700");
  // Value — row 4
  e.append("text").attr("class","nv").attr("x",NW/2).attr("y",68)
   .attr("text-anchor","middle").attr("font-size","12px").attr("font-weight","700").attr("font-family","monospace");
  // FT label — row 5
  e.append("text").attr("class","nfl").attr("x",NW/2).attr("y",84)
   .attr("text-anchor","middle").attr("font-size","8px").attr("fill","#7e57c2");
  // Pin/icon top-right corner
  e.append("text").attr("class","nfx").attr("x",NW-6).attr("y",13).attr("text-anchor","end")
   .attr("font-size","8px").attr("fill","#e94560");
  const all=e.merge(s);

  all.on("click",(ev,d)=>{
    ev.stopPropagation();
    if(ev.shiftKey||ev.ctrlKey||ev.metaKey){toggleMultiSel(d.id,ev);return;}
    selectNode(d.id);
  })
  .on("dblclick",(ev,d)=>{
    ev.stopPropagation();
    if(d.children?.length){collapsed.has(d.id)?collapsed.delete(d.id):collapsed.add(d.id);refresh();}
  })
  // Drag works whether force is on or off
  .call(d3.drag()
    .on("start",(ev,d)=>{
      ev.sourceEvent.stopPropagation();
      d.fx=d.x; d.fy=d.y;
      if(forceOn) sim.alphaTarget(0.05).restart();
    })
    .on("drag",(ev,d)=>{
      d.fx=ev.x; d.fy=ev.y;
      uP[d.id]={x:ev.x,y:ev.y,manual:true};
      if(forceOn) sim.alphaTarget(0.05).restart();
      else tick();
    })
    .on("end",(ev,d)=>{
      // Keep fx/fy so node stays where dropped (force off = sticky)
      if(forceOn){d.fx=null;d.fy=null;}
      else{d.fx=d.x;d.fy=d.y;}
      savePositions();
    })
  );

  all.select("rect.nb")
    .attr("stroke",d=>d.isRoot?"#ffffff44":d.isDuplicate?"#4fc3f7":d.isPinned?"#e94560":d.color)
    .attr("stroke-width",d=>selId===d.id?3:d.isDuplicate?2.5:d.isPinned?2.5:1.5)
    .attr("stroke-dasharray",d=>selId===d.id?null:d.isRoot?"4,4":d.isDuplicate?"6,3":d.isPinned?"6,3":null)
    .attr("filter",d=>selId===d.id?`drop-shadow(0 0 12px ${d.color}88)`:null)
    .attr("fill",d=>`${d.color}18`);

  all.select("text.nt").text(d=>d.isGroup?"GROUP":d.type).attr("fill",d=>d.color);
  all.select("text.ngt").text(d=>d.gate).attr("fill",d=>d.gate==="AND"?"#ffb74d":"#4fc3f7");

  // ID prefix colored, number grey
  all.select("text.nidp").each(function(d){
    const nid=d.nodeId&&d.nodeId!==d.id?d.nodeId:"";
    if(!nid){d3.select(this).text("");return;}
    const m=nid.match(/^([A-Za-z]+-?)(\d+.*)$/);
    d3.select(this).text(m?m[1]:nid).attr("fill",d.color);
  });
  all.select("text.nidn").each(function(d){
    const nid=d.nodeId&&d.nodeId!==d.id?d.nodeId:"";
    if(!nid){d3.select(this).text("").attr("x",8);return;}
    const m=nid.match(/^([A-Za-z]+-?)(\d+.*)$/);
    if(!m){d3.select(this).text("").attr("x",8);return;}
    d3.select(this).text(m[2]).attr("x",8+m[1].length*6.2);
  });
  all.select("text.nn").text(d=>{const nm=d.name||"";return nm.length>22?nm.slice(0,21)+"…":nm;}).attr("fill",d=>d.color);
  all.select("text.nv").text(d=>d.isPinned?(d.fixedVal||d.value)+" 📌":d.value).attr("fill",d=>d.isPinned?"#e94560":d.color);
  all.select("text.nfl").text(d=>d.ftLabel?`[${d.ftLabel}]`:"");
  all.select("text.nfx").text(d=>d.isPinned?"📌":d.isRoot?"⬡":d.isDuplicate?"◈":"");
  s.exit().remove();
}

function tick(){
  ng.selectAll("g.nd").attr("transform",d=>`translate(${(d.x||0)-NW/2},${(d.y||0)-NH/2})`);
  // Lines: source=child (bottom), target=parent (top)
  // Start at top of child node, end at bottom of parent node
  lg.selectAll("path.lk").attr("d",d=>{
    const cx=d.source.x||0,cy=(d.source.y||0)-NH/2;   // top of child
    const px=d.target.x||0,py=(d.target.y||0)+NH/2;   // bottom of parent
    const my=(cy+py)/2;
    return `M${cx},${cy} C${cx},${my} ${px},${my} ${px},${py}`;
  });
  // Gate symbols at bottom of parent node
  gg.selectAll("g.gsym").attr("transform",d=>{
    const px=d.n.x||0, py=(d.n.y||0)+NH/2+12;
    return `translate(${px},${py})`;
  });
}

function updateLanes(){
  const lc=document.getElementById("lanes"); lc.innerHTML="";
  const t=getT(),h=wrap.getBoundingClientRect().height;
  const rows=new Map();
  sN.forEach(n=>{const r=getNodeRow(n);if(!rows.has(r))rows.set(r,[]);rows.get(r).push(n);});
  const extraLbls={"1.5":"SHARED FF/GROUP"};
  rows.forEach((_,r)=>{
    const cy=t.k*(layY({_midRow:r,type:"HAZARD"}))+t.y;
    if(cy<-40||cy>h+40) return;
    const lbl=LLABELS[String(r)]||extraLbls[String(r)]||"";
    const col=LCOLORS[String(r)]||"#888";
    const div=document.createElement("div");
    div.className="lb"; div.style.top=(cy-52)+"px";
    const bar=document.createElement("div"); bar.className="ls"; bar.style.background=col;
    const txt=document.createElement("div"); txt.className="lt"; txt.textContent=lbl; txt.style.color=col;
    div.appendChild(bar); div.appendChild(txt); lc.appendChild(div);
  });
}

// ── Hierarchical column layout ─────────────────────────────────────────────
function computeRTLayout(reset){
  const sfNodes=sN.filter(n=>n.type==="SF");
  const totalSFs=sfNodes.length||1;
  const nodeSpacing=NW+HG+80;
  const canvasW=Math.max(totalSFs*nodeSpacing+400,1600);
  const startX=220;
  const colStep=(canvasW-startX*2)/(totalSFs>1?totalSFs-1:1);

  const sfColX={};
  sfNodes.forEach((sf,i)=>{
    const cx=startX+i*(totalSFs>1?colStep:0);
    sfColX[sf.id]=cx; sf.bx=cx;
    if(reset||!uP[sf.id]?.manual){sf.x=cx;sf.fx=null;sf.fy=null;}
  });

  // Map descendants to SF columns
  const nodeSFCols={};
  sfNodes.forEach(sf=>{
    subIds(sf.id).forEach(nid=>{
      if(!nodeSFCols[nid]) nodeSFCols[nid]=[];
      nodeSFCols[nid].push(sfColX[sf.id]);
    });
  });

  // Bucket nodes by (centerX, row) and spread them
  const buckets={};
  sN.forEach(n=>{
    if(n.type==="SF"||n.type==="HAZARD") return;
    const cols=nodeSFCols[n.id];
    const cx=cols&&cols.length?cols.reduce((a,b)=>a+b,0)/cols.length:canvasW+200;
    n.bx=cx;
    const row=getNodeRow(n);
    const key=`${Math.round(cx/50)*50}_${row}`;
    if(!buckets[key]) buckets[key]=[];
    buckets[key].push(n);
  });
  Object.values(buckets).forEach(bucket=>{
    if(bucket.length<=1) return;
    const spacing=NW+55;
    const center=bucket[0].bx;
    bucket.forEach((n,i)=>{
      n.bx=center+(i-(bucket.length-1)/2)*spacing;
      if(reset||!uP[n.id]?.manual){n.x=n.bx;n.fx=null;n.fy=null;}
    });
  });

  // Assign bx→x for nodes not in buckets
  sN.forEach(n=>{
    if(n.type==="SF"||n.type==="HAZARD") return;
    if(reset||!uP[n.id]?.manual){n.x=n.bx||canvasW+200;n.fx=null;n.fy=null;}
  });

  // HAZARDs centred over SF children
  sN.filter(n=>n.type==="HAZARD").forEach(n=>{
    const cs=sN.filter(c=>c.type==="SF"&&c.parents?.includes(n.id));
    n.bx=cs.length?cs.reduce((a,c)=>a+c.bx,0)/cs.length:canvasW/2;
    if(reset||!uP[n.id]?.manual){n.x=n.bx;n.fx=null;n.fy=null;}
  });

  // Y by row
  sN.forEach(n=>{
    const ty=layY(n);
    if(reset||!uP[n.id]?.manual){n.y=ty;n.fy=null;}
  });
}

function doColumnLayout(reset){
  Object.keys(uP).forEach(id=>{if(uP[id])uP[id].manual=false;});
  computeRTLayout(true);
  sN.forEach(n=>{n.vx=0;n.vy=0;
    // Release sticky pins for reset
    n.fx=null;n.fy=null;
  });
  tick(); updateLanes();
  setTimeout(doFit,80);
}
function doFit(){
  if(!sN.length) return;
  const xs=sN.map(n=>n.x||0),ys=sN.map(n=>n.y||0);
  const minX=Math.min(...xs)-NW,maxX=Math.max(...xs)+NW;
  const minY=Math.min(...ys)-NH,maxY=Math.max(...ys)+NH;
  const cw=wrap.getBoundingClientRect();
  const pad=50,k=Math.min(.88,(cw.width-pad*2)/(maxX-minX+1),(cw.height-pad*2)/(maxY-minY+1));
  const tx=cw.width/2-k*(minX+maxX)/2,ty=cw.height/2-k*(minY+maxY)/2;
  svg.transition().duration(600).call(zb.transform,d3.zoomIdentity.translate(tx,ty).scale(k));
  setTimeout(updateLanes,620);
}
function zBy(d){
  const t=getT(),cw=wrap.getBoundingClientRect();
  svg.transition().duration(200).call(zb.scaleBy,1+d,[cw.width/2,cw.height/2]);
  setTimeout(updateLanes,220);
}
function toggleForce(){
  const n=sN.find(n=>n.id===selId&&n.type==="SF");
  if(n){forceSub=subIds(n.id);forceOn=true;}
  else{forceSub=null;forceOn=!forceOn;}
  document.getElementById("bfrc").classList.toggle("on",forceOn);
  const fst=document.getElementById("fst");
  if(forceOn){
    const ifc=n?countIFs(n.id):sN.filter(x=>x.type==="IF").length;
    fst.textContent=`⚡ ${n?n.name.slice(0,16):"ALL"}  ·  ${ifc} IF`;
    fst.classList.add("show");
    // Release fx/fy so force can move nodes
    sN.forEach(x=>{x.fx=null;x.fy=null;});
    sim.nodes(forceSub?sN.filter(n=>forceSub.has(n.id)):sN);
    sim.force("link").links(forceSub?sL.filter(l=>forceSub.has(l.source.id)&&forceSub.has(l.target.id)):sL);
    sim.alpha(.5).restart();
  } else {
    fst.classList.remove("show");
    sim.stop();
    // Re-pin all nodes to current positions
    sN.forEach(n=>{n.fx=n.x;n.fy=n.y;});
  }
}
function clearHL(){
  ng.selectAll("g.nd").attr("opacity",d=>d.isRoot?0.7:1);
  lg.selectAll("path.lk").attr("opacity",1);
  gg.selectAll("g.gsym").attr("opacity",1);
}

// ── Right Panel node inspector / editor ───────────────────────────────────
function selectNode(id){
  selId=id;
  const n=NM[id]; if(!n) return;
  sendSelNode(id);
  // Highlight connected nodes
  clearHL();
  ng.selectAll("g.nd").each(function(x){
    const inPath=x.parents?.includes(id)||x.children?.includes(id)||x.id===id;
    d3.select(this).attr("opacity",inPath?1:0.15);
  });
  lg.selectAll("path.lk").attr("opacity",l=>{
    return (l.parentId===id||l.childId===id)?1:0.06;
  });
  gg.selectAll("g.gsym").attr("opacity",d=>d.n.id===id?1:0.06);
  // Update selected node stroke
  ng.selectAll("rect.nb")
    .attr("stroke",d=>selId===d.id?"#fff":d.isRoot?"#ffffff44":d.isDuplicate?"#4fc3f7":d.isPinned?"#e94560":d.color)
    .attr("stroke-width",d=>selId===d.id?3:d.isDuplicate?2.5:d.isPinned?2.5:1.5)
    .attr("filter",d=>selId===d.id?`drop-shadow(0 0 14px ${d.color}aa)`:null);
  // Populate right panel
  openRP(id);
}
function openRP(id){
  const n=NM[id]; if(!n) return;
  const rp=document.getElementById("rp");
  rp.classList.add("open");
  // Header
  document.getElementById("rp-name").textContent=n.name||"";
  document.getElementById("rp-name").style.color=n.color;
  document.getElementById("rp-badge").innerHTML=
    `<code style="background:#1a1a1a;color:${n.color};font-size:9px;padding:1px 6px;border-radius:3px;">${n.nodeId||id}</code>`+
    `<code style="background:#1a1a1a;color:#555;font-size:9px;padding:1px 6px;border-radius:3px;margin-left:3px;">${n.type}</code>`+
    (n.shared?`<span style="background:#f5c51822;color:#f5c518;font-size:8px;padding:1px 5px;border-radius:3px;margin-left:3px;border:1px solid #f5c51844;">SHARED</span>`:"");
  document.getElementById("rp-type").textContent=n.isGroup?"GROUP":n.type;
  document.getElementById("rp-type").style.color=n.color;
  document.getElementById("rp-gate-disp").textContent=n.gate;
  document.getElementById("rp-gate-disp").style.color=n.gate==="AND"?"#ffb74d":"#4fc3f7";
  document.getElementById("rp-value").textContent=n.isPinned?(n.fixedVal||n.value)+" 📌":n.value;
  document.getElementById("rp-value").style.color=n.isPinned?"#e94560":n.color;
  document.getElementById("rp-nid").textContent=n.nodeId||id;
  document.getElementById("rp-nid").style.color="#aaa";
  // IF count
  const ifc=countIFs(id);
  document.getElementById("rp-ifcount").textContent=ifc+" IF";
  // Parents
  const phtml=(n.pnames||[]).map((nm,i)=>{
    const pid=(n.parents||[])[i];
    return `<span class="rp-chip" style="background:#1a1a1a;border:1px solid #333;color:#bbb;"
      onclick="selectNode('${pid}')">${nm.slice(0,24)}</span>`;
  }).join("")||`<span style="color:#333;font-size:9px;">(top event)</span>`;
  document.getElementById("rp-parents").innerHTML=phtml;
  // Children
  const chtml=(n.cnames||[]).map((nm,i)=>{
    const chid=(n.children||[])[i];
    return `<span class="rp-chip" style="background:#1a1a1a;border:1px solid #333;color:#bbb;"
      onclick="selectNode('${chid}')">${nm.slice(0,24)}</span>`;
  }).join("")||`<span style="color:#333;font-size:9px;">(leaf node)</span>`;
  document.getElementById("rp-children").innerHTML=chtml;
  // Edit fields
  document.getElementById("rp-edit-name").value=n.name||"";
  document.getElementById("rp-edit-gate").value=n.gate||"OR";
  document.getElementById("rp-edit-nodeid").value=n.nodeId||"";
  document.getElementById("rp-edit-ftlabel").value=n.ftLabel||"";
  document.getElementById("rp-edit-fixed").value=n.fixedVal||"";
  document.getElementById("rp-edit-target").value=n.targetValue||"";
  // Hide delete for HAZARD
  document.getElementById("rp-del-btn").style.display=n.type==="HAZARD"?"none":"block";
  document.getElementById("rp-save-msg").style.display="none";
}
function closeRP(){
  document.getElementById("rp").classList.remove("open");
  clearHL();
  ng.selectAll("rect.nb")
    .attr("stroke",d=>d.isRoot?"#ffffff44":d.isDuplicate?"#4fc3f7":d.isPinned?"#e94560":d.color)
    .attr("stroke-width",d=>d.isDuplicate?2.5:d.isPinned?2.5:1.5)
    .attr("filter",null);
  selId=null;
  try{window.parent.postMessage(JSON.stringify({type:"fta_selnode",data:null}),"*");}catch(e){}
}
function rpSave(){
  if(!selId) return;
  const name=document.getElementById("rp-edit-name").value.trim();
  const gate=document.getElementById("rp-edit-gate").value;
  const nodeId=document.getElementById("rp-edit-nodeid").value.trim();
  const ftLabel=document.getElementById("rp-edit-ftlabel").value.trim();
  const fixedStr=document.getElementById("rp-edit-fixed").value.trim();
  const targetStr=document.getElementById("rp-edit-target").value.trim();
  const fixedVal=fixedStr?parseFloat(fixedStr):null;
  const targetVal=targetStr?parseFloat(targetStr):null;
  try{
    window.parent.postMessage(JSON.stringify({
      type:"fta_edit_node",
      data:{id:selId,name,gate,nodeId,ftLabel,fixedVal,targetVal}
    }),"*");
    // Update local NM for immediate visual refresh
    if(NM[selId]){
      NM[selId].name=name; NM[selId].gate=gate;
      NM[selId].nodeId=nodeId; NM[selId].ftLabel=ftLabel;
      if(fixedVal!=null){NM[selId].fixedVal=String(fixedVal);NM[selId].isPinned=true;}
      else{NM[selId].fixedVal=null;NM[selId].isPinned=false;}
    }
    const snode=sN.find(n=>n.id===selId);
    if(snode){
      snode.name=name;snode.gate=gate;snode.nodeId=nodeId;snode.ftLabel=ftLabel;
      if(fixedVal!=null){snode.fixedVal=String(fixedVal);snode.isPinned=true;}
      else{snode.fixedVal=null;snode.isPinned=false;}
    }
    drawNodes();
    document.getElementById("rp-save-msg").style.display="block";
    setTimeout(()=>{document.getElementById("rp-save-msg").style.display="none";},2000);
  }catch(ex){}
}
function rpDelete(){
  if(!selId) return;
  if(!confirm("Delete this node and its orphaned children?")) return;
  try{
    window.parent.postMessage(JSON.stringify({type:"fta_delete_node",data:{id:selId}}),"*");
    closeRP();
  }catch(e){}
}
function doSearch(q){
  ng.selectAll("rect.nb").attr("filter",d=>selId===d.id?`drop-shadow(0 0 14px ${d.color}aa)`:null);
  sM=[];sI=0;
  if(!q.trim()){document.getElementById("si").textContent="";return;}
  const lq=q.toLowerCase();
  RNODES.forEach(n=>{if([n.name,n.type,n.value,n.nodeId||""].some(v=>v.toLowerCase().includes(lq)))sM.push(n.id);});
  document.getElementById("si").textContent=sM.length?sM.length+" found":"0";
  ng.selectAll("g.nd").filter(d=>sM.includes(d.id)).select("rect.nb").attr("filter","drop-shadow(0 0 10px #f5c518)");
  if(sM.length)panTo(sM[0]);
}
function sNext(){if(!sM.length)return;sI=(sI+1)%sM.length;panTo(sM[sI]);document.getElementById("si").textContent=`${sI+1}/${sM.length}`;}
function sPrev(){if(!sM.length)return;sI=(sI-1+sM.length)%sM.length;panTo(sM[sI]);document.getElementById("si").textContent=`${sI+1}/${sM.length}`;}
function panTo(id){
  const n=sN.find(s=>s.id===id);if(!n)return;
  const cw=wrap.getBoundingClientRect();
  const k=Math.min(getT().k,0.55);
  svg.transition().duration(500).call(zb.transform,d3.zoomIdentity.translate(cw.width/2-k*(n.x||0),cw.height/2.5-k*(n.y||0)).scale(k));
  setTimeout(updateLanes,520);
}
function toggleMultiSel(id){
  if(multiSel.has(id)) multiSel.delete(id); else multiSel.add(id);
  updateMultiSelUI(); sendMultiSel();
}
function updateMultiSelUI(){
  const msp=document.getElementById("msp");
  const count=multiSel.size;
  document.getElementById("msc").textContent=count;
  const list=document.getElementById("mslist"); list.innerHTML="";
  multiSel.forEach(id=>{
    const n=NM[id]; if(!n) return;
    const chip=document.createElement("span");
    chip.style.cssText=`background:#0a1a2e;border:1px solid ${n.color};color:${n.color};font-size:9px;padding:1px 6px;border-radius:10px;font-family:monospace;font-weight:700;cursor:pointer;`;
    chip.textContent=(n.nodeId||n.id)+" "+n.name.slice(0,16);
    chip.onclick=()=>{multiSel.delete(id);updateMultiSelUI();sendMultiSel();};
    list.appendChild(chip);
  });
  msp.style.display=count>0?"block":"none";
  ng.selectAll("g.nd").each(function(d){
    const isSel=multiSel.has(d.id);
    d3.select(this).attr("opacity",count>0?(isSel?1:0.2):d.isRoot?0.7:1)
      .select("rect.nb")
      .attr("stroke",isSel?"#e94560":d.isRoot?"#ffffff44":d.isDuplicate?"#4fc3f7":d.isPinned?"#e94560":d.color)
      .attr("stroke-width",isSel?3.5:d.isDuplicate?2.5:d.isPinned?2.5:1.5)
      .attr("stroke-dasharray",isSel?null:d.isRoot?"4,4":d.isDuplicate?"6,3":d.isPinned?"6,3":null)
      .attr("filter",isSel?"drop-shadow(0 0 10px #e9456088)":null);
  });
  gg.selectAll("g.gsym").attr("opacity",count>0?0.15:1);
}
function clearMultiSel(){
  multiSel.clear(); updateMultiSelUI(); sendMultiSel();
  ng.selectAll("g.nd").each(function(d){
    d3.select(this).attr("opacity",d.isRoot?0.7:1)
      .select("rect.nb")
      .attr("stroke",d.isRoot?"#ffffff44":d.isDuplicate?"#4fc3f7":d.isPinned?"#e94560":d.color)
      .attr("stroke-width",d.isDuplicate?2.5:d.isPinned?2.5:1.5)
      .attr("stroke-dasharray",d.isRoot?"4,4":d.isDuplicate?"6,3":d.isPinned?"6,3":null)
      .attr("filter",null);
  });
  gg.selectAll("g.gsym").attr("opacity",1);
}
function sendMultiSel(){
  try{window.parent.postMessage(JSON.stringify({type:"fta_multisel",data:Array.from(multiSel).map(id=>{const n=NM[id]||{};return {id,name:n.name||id,nodeId:n.nodeId||id,type:n.type||"",color:n.color||"#888",parents:n.parents||[],children:n.children||[]};})}),"*");}catch(e){}
}
function sendSelNode(id){
  try{const n=NM[id]||{};window.parent.postMessage(JSON.stringify({type:"fta_selnode",data:{id,name:n.name||id,nodeId:n.nodeId||id,type:n.type||"",color:n.color||"#888",gate:n.gate||"OR",value:n.value||"-",isPinned:n.isPinned||false,fixedVal:n.fixedVal||null,parents:n.parents||[],pnames:n.pnames||[],children:n.children||[],cnames:n.cnames||[]}}),"*");}catch(e){}
}
function savePositions(){
  const pos={};
  sN.forEach(n=>{if(n.x!=null&&n.y!=null)pos[n.id]={x:Math.round(n.x),y:Math.round(n.y),manual:!!(uP[n.id]?.manual)};});
  try{window.parent.postMessage(JSON.stringify({type:"fta_pos",data:pos}),"*");}catch(e){}
}
setInterval(savePositions,15000);
document.addEventListener("pointerup",()=>setTimeout(savePositions,300));

refresh();
const hasSavedPos=Object.keys(IPOS).length>0;
if(hasSavedPos){
  Object.entries(IPOS).forEach(([id,p])=>{if(p&&p.x!=null){uP[id]={x:p.x,y:p.y,manual:true};const n=sN.find(x=>x.id===id);if(n){n.x=p.x;n.y=p.y;n.fx=p.x;n.fy=p.y;}}});
  computeRTLayout(false);tick();updateLanes();
  setTimeout(doFit,80);
} else {doColumnLayout(true);}
if(FOCUSID&&!hasSavedPos) setTimeout(()=>panTo(FOCUSID),900);
window.addEventListener("message",function(e){
  try{
    const d=typeof e.data==="string"?JSON.parse(e.data):e.data;
    if(!d) return;
    if(d.type==="fta_restore_pos"){
      let changed=false;
      Object.entries(d.data).forEach(([id,p])=>{if(p&&p.x!=null&&!uP[id]?.manual){uP[id]={x:p.x,y:p.y,manual:p.manual||false};changed=true;}});
      if(changed){computeRTLayout(false);tick();updateLanes();}
    }
  }catch(err){}
});
</script></body></html>"""

    html = html.replace("__NODES__", nodes_js)
    html = html.replace("__LINKS__", links_js)
    html = html.replace("__IPOS__",  init_pos_js)
    html = html.replace("__FOCUS__", focus_js)
    html = html.replace("__LLABELS__", level_label_js)
    html = html.replace("__LCOLORS__", level_color_js)
    return html

def build_hierarchy_rows(nodes, filter_hazard_id=None):
    by_id = {n["id"]: n for n in nodes}
    rows, visited = [], set()
    def walk(nid, depth):
        is_ref = nid in visited
        if not is_ref: visited.add(nid)
        node = by_id.get(nid)
        if not node: return
        rows.append({"node": node, "depth": depth, "ref": is_ref})
        if not is_ref:
            for child in [n for n in nodes if nid in (n.get("parentIds") or [])]:
                walk(child["id"], depth + 1)
    starts = ([n for n in nodes if n["type"]=="HAZARD" and n["id"]==filter_hazard_id]
              if filter_hazard_id else [n for n in nodes if n["type"]=="HAZARD"])
    for h in starts: walk(h["id"], 0)
    return rows

# ── Session state ─────────────────────────────────────────────────────────
DEFS = {"nodes":[],"save_status":"idle","save_msg":"","gist_loaded":False,
        "active_file":"my_tree.json","file_list":[],"selected_id":None,
        "tree_filter":"ALL",
        "nodes_since_calc": 0,
        "pending_node_names": [],
        "nodes_hash": "",
        "multisel_ids": [],
        "selected_node": None,
        "_pending_positions": {},
        "tree_state": {
            "scale": 1.0, "tx": 0, "ty": 0,
            "collapsed": [],
            "positions": {},
            "focus_id": None,
        }
        }
for k,v in DEFS.items():
    if k not in st.session_state: st.session_state[k] = v

def get_secret(k):
    try: return st.secrets[k]
    except: return os.environ.get(k,"")

GITHUB_TOKEN = get_secret("GITHUB_TOKEN")
GIST_ID      = get_secret("GIST_ID")
configured   = bool(GITHUB_TOKEN and GIST_ID)

if configured and not st.session_state.gist_loaded:
    with st.spinner("Loading from Gist..."):
        st.session_state.file_list = list_gist_files(GITHUB_TOKEN, GIST_ID)
        af = st.session_state.active_file
        if af in st.session_state.file_list:
            _loaded = load_gist_file(GITHUB_TOKEN, GIST_ID, af)
        elif st.session_state.file_list:
            named = [f for f in st.session_state.file_list if is_named(f)]
            if named:
                st.session_state.active_file = named[0]
                _loaded = load_gist_file(GITHUB_TOKEN, GIST_ID, named[0])
            else:
                _loaded = []
        else:
            _loaded = []
        st.session_state.nodes = _loaded
        _restored_pos = extract_positions(_loaded)
        if _restored_pos:
            st.session_state["_pending_positions"] = _restored_pos
            st.session_state.tree_state["positions"] = _restored_pos
        st.session_state.gist_loaded = True
        st.session_state.save_status = "loaded"
        st.session_state.save_msg = f"Loaded '{st.session_state.active_file}'"

def save_current(nodes=None, filename=None, status_label=None):
    if nodes   is None: nodes    = st.session_state.nodes
    if filename is None: filename = st.session_state.active_file
    if configured:
        ok = save_gist_file(GITHUB_TOKEN, GIST_ID, filename, nodes)
        st.session_state.save_status = "saved" if ok else "error"
        st.session_state.save_msg = (status_label or
            f"Saved '{filename}' at {datetime.now().strftime('%H:%M:%S')}") if ok else "Save failed"
        st.session_state.file_list = list_gist_files(GITHUB_TOKEN, GIST_ID)
        return ok
    st.session_state.save_status = "no_config"
    st.session_state.save_msg = "Gist not configured"
    return False

def set_nodes(n, recalc=False):
    if recalc:
        n = recalculate(n)
        st.session_state.nodes_since_calc = 0
        st.session_state.pending_node_names = []
        st.session_state.nodes_hash = ""
    pending_pos = st.session_state.get("_pending_positions", {})
    if pending_pos:
        n = inject_positions(n, pending_pos)
    st.session_state.nodes = n
    save_current(n)

# ── CSS ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&display=swap');
html,body,[class*="css"]{font-family:'JetBrains Mono',monospace!important;background:#0d0d0d!important;color:#e0e0e0!important;}
.stApp{background:#0d0d0d!important;}
section[data-testid="stSidebar"]{background:#111!important;border-right:1px solid #222!important;}
.stButton>button{font-family:'JetBrains Mono',monospace!important;font-weight:700!important;letter-spacing:1px!important;}
.stTabs [data-baseweb="tab"]{font-family:'JetBrains Mono',monospace!important;font-size:10px!important;}
</style>""", unsafe_allow_html=True)

# ── Data ──────────────────────────────────────────────────────────────────
nodes    = st.session_state.nodes
hazards  = [n for n in nodes if n["type"] == "HAZARD"]
by_id    = {n["id"]: n for n in nodes}
by_level = {lvl: [n for n in nodes if n["type"] == lvl] for lvl in DISPLAY_ORDER}

# ── Header ────────────────────────────────────────────────────────────────
sc = st.session_state.save_status
sc_color = {"saved":"#4caf7d","loaded":"#4caf7d","error":"#ff4d4d","no_config":"#f5c518","idle":"#888"}.get(sc,"#888")
sc_icon  = {"saved":"✓","loaded":"↓","error":"✗","no_config":"!","idle":"○"}.get(sc,"○")
st.markdown(f"""
<div style="background:linear-gradient(90deg,#1a1a2e,#16213e,#0f3460);
            border-bottom:2px solid #e94560;padding:11px 20px;
            margin:-1rem -1rem 1rem -1rem;
            display:flex;justify-content:space-between;align-items:center;">
  <div>
    <div style="font-size:19px;font-weight:700;letter-spacing:2px;color:#e94560;">
      ⚠ FTA REVERSE ENGINEER
    </div>
    <div style="font-size:9px;color:#888;letter-spacing:3px;margin-top:1px;">
      FAULT TREE ANALYSIS · TOP-DOWN DISTRIBUTION · {len(nodes)} nodes · {len(hazards)} hazard(s){f" · {sum(1 for n in nodes if n.get('fixedValue') is not None)} pinned 📌" if any(n.get('fixedValue') is not None for n in nodes) else ""}
    </div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:11px;color:{sc_color};font-weight:700;">{sc_icon} {st.session_state.save_msg or "Ready"}</div>
    <div style="font-size:9px;color:#555;margin-top:1px;">Active: <span style="color:#aaa;">{st.session_state.active_file}</span></div>
  </div>
</div>""", unsafe_allow_html=True)

if not configured:
    st.warning("Gist not configured — data resets on refresh. Add GITHUB_TOKEN + GIST_ID to Streamlit secrets.")

# ── Main tabs ─────────────────────────────────────────────────────────────
tab_tree, tab_verify, tab_hier, tab_data, tab_search = st.tabs([
    "🌳 TREE", "🔬 VERIFY", "📋 HIERARCHY", "📊 DATA", "🔍 SEARCH"
])

# ── Sidebar (unchanged — render_sidebar inline) ───────────────────────────
@st.fragment
def render_sidebar():
    nodes  = st.session_state.nodes
    by_id  = {n["id"]: n for n in nodes}
    hazards = [n for n in nodes if n["type"] == "HAZARD"]

    # ── Single-node selection panel (from canvas click) ──────────────────
    sel_node_data = st.session_state.get("selected_node")
    if sel_node_data and isinstance(sel_node_data, dict):
        snd    = sel_node_data
        sn_id  = snd.get("id")
        sn_obj = by_id.get(sn_id)
        c      = LEVEL_COLORS.get(snd.get("type",""), "#888")
        st.markdown(
            f"<div style='background:#0d1a0d;border:2px solid {c};border-radius:8px;"
            f"padding:10px 14px;margin-bottom:8px;'>"
            f"<div style='font-size:8px;color:{c};font-weight:700;letter-spacing:2px;margin-bottom:5px;'>&#11044; SELECTED NODE</div>"
            f"<div style='font-weight:700;font-size:12px;color:#ddd;margin-bottom:2px;'>{esc(snd.get('name','?'))}</div>"
            f"<div style='font-family:monospace;font-size:9px;color:{c};'>{esc(snd.get('nodeId',''))} &nbsp;·&nbsp; "
            f"{snd.get('type','')} &nbsp;·&nbsp; GATE:{snd.get('gate','')}</div>"
            f"<div style='font-size:11px;color:{c};font-family:monospace;font-weight:700;margin-top:4px;'>{snd.get('value','-')}"
            f"{'&nbsp;&#128204;' if snd.get('isPinned') else ''}</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        if sn_obj:
            _ac1, _ac2, _ac3 = st.columns(3)
            with _ac1:
                if st.button("&#128065; Info", key="sn_info_btn", use_container_width=True,
                             help="Jump to this node in the DATA tab"):
                    st.session_state["selected_node"] = None
                    nid_str  = snd.get('nodeId','')
                    nm_str   = snd.get('name','')
                    type_str = snd.get('type','')
                    gate_str = snd.get('gate','')
                    val_str  = snd.get('value','-')
                    st.info(f"Node: {nid_str} — {nm_str}\nType: {type_str} | Gate: {gate_str}\nValue: {val_str}")
            with _ac2:
                if st.button("&#9998; Edit", key="sn_edit_btn", use_container_width=True,
                             help="Pre-select this node in the EDIT tab"):
                    st.session_state["_sidebar_edit_target"] = sn_id
                    st.session_state["selected_node"] = None
                    st.rerun()
            with _ac3:
                if sn_obj.get("type") != "HAZARD":
                    if st.button("&#128465; Delete", key="sn_del_btn", use_container_width=True,
                                 help="Delete this node and its orphaned children"):
                        pre_roots = {n["id"] for n in nodes if n["type"] != "HAZARD" and not n.get("parentIds")}
                        tmp = [dict(n) for n in nodes if n["id"] != sn_id]
                        for n in tmp:
                            if sn_id in (n.get("parentIds") or []):
                                n["parentIds"] = [p for p in n["parentIds"] if p != sn_id]
                        changed = True
                        while changed:
                            changed = False
                            orphans = {n["id"] for n in tmp if n["type"] != "HAZARD" and not n.get("parentIds") and n["id"] not in pre_roots}
                            if orphans:
                                tmp = [n for n in tmp if n["id"] not in orphans]
                                for n in tmp:
                                    before = len(n.get("parentIds") or [])
                                    n["parentIds"] = [p for p in (n.get("parentIds") or []) if p not in orphans]
                                    if len(n.get("parentIds") or []) != before: changed = True
                        st.session_state["selected_node"] = None
                        st.session_state.nodes_since_calc += 1
                        set_nodes(tmp)
                        st.rerun()
                else:
                    st.markdown("<div style='font-size:8px;color:#555;padding-top:6px;'>HAZARD — cannot delete</div>",
                                unsafe_allow_html=True)
        st.markdown("<hr style='border-color:#1a1a1a;margin:6px 0;'>", unsafe_allow_html=True)

    sel_ids   = st.session_state.get("multisel_ids", [])
    sel_nodes = [by_id[i] for i in sel_ids if i in by_id]

    if sel_nodes:
        st.markdown(f"""
        <div style="background:#080f1a;border:2px solid #4fc3f7;border-radius:8px;
                    padding:10px 14px;margin-bottom:8px;">
          <div style="font-size:8px;color:#4fc3f7;font-weight:700;letter-spacing:2px;margin-bottom:4px;">
            SELECTION — {len(sel_nodes)} NODES
          </div>""", unsafe_allow_html=True)
        for sn in sel_nodes[:5]:
            c = LEVEL_COLORS.get(sn["type"],"#888")
            st.markdown(
                f"<div style='font-size:9px;color:{c};font-family:monospace;padding:1px 0;'>"
                f"{'◈ ' if len(sn.get('parentIds') or [])>1 else ''}"
                f"{esc(sn.get('nodeId',sn['id']))} {esc(sn['name'][:28])}</div>",
                unsafe_allow_html=True)
        if len(sel_nodes) > 5:
            st.markdown(f"<div style='font-size:9px;color:#555;'>+ {len(sel_nodes)-5} more</div>",
                        unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with st.expander("📁 FILES", expanded=False):
        if configured:
            fl = st.session_state.file_list
            named_files = [f for f in fl if is_named(f)]
            if named_files:
                cur = st.session_state.active_file
                idx = named_files.index(cur) if cur in named_files else 0
                chosen = st.selectbox("File", named_files, index=idx, key="file_sel",
                                      label_visibility="collapsed")
                if chosen != st.session_state.active_file:
                    st.session_state.active_file = chosen
                    _l = load_gist_file(GITHUB_TOKEN, GIST_ID, chosen)
                    st.session_state.nodes = _l
                    _rp = extract_positions(_l)
                    if _rp:
                        st.session_state["_pending_positions"] = _rp
                        st.session_state.tree_state["positions"] = _rp
                    st.rerun()
            new_fname = st.text_input("New file name", placeholder="project_v2.json", key="new_fname")
            if st.button("＋ Create", use_container_width=True) and new_fname.strip():
                fn = new_fname.strip()
                if not fn.endswith(".json"): fn += ".json"
                save_gist_file(GITHUB_TOKEN, GIST_ID, fn, [])
                st.session_state.active_file = fn
                st.session_state.nodes = []
                st.session_state.file_list = list_gist_files(GITHUB_TOKEN, GIST_ID)
                st.rerun()
        else:
            st.markdown("<div style='font-size:9px;color:#555;'>Configure Gist to use files.</div>",
                        unsafe_allow_html=True)

    with st.expander("📥 IMPORT / EXPORT", expanded=False):
        # ── Export scope notice ──────────────────────────────────────────
        n_hazards = len([n for n in st.session_state.nodes if n["type"] == "HAZARD"])
        st.markdown(
            f"<div style='background:#0a1a0a;border:1px solid #4caf7d44;border-radius:5px;"
            f"padding:6px 10px;margin-bottom:8px;'>"
            f"<div style='font-size:8px;color:#4caf7d;font-weight:700;letter-spacing:1px;margin-bottom:3px;'>&#9432; EXPORT SCOPE</div>"
            f"<div style='font-size:9px;color:#888;line-height:1.5;'>"
            f"All exports contain the <b style='color:#ddd;'>entire fault tree</b> — all {len(st.session_state.nodes)} nodes "
            f"across all {n_hazards} hazard(s). There is no per-hazard export; "
            f"the tree is one connected dataset. To get per-hazard data, use the "
            f"<b style='color:#ddd;'>HIERARCHY</b> tab filter and copy manually."
            f"</div></div>",
            unsafe_allow_html=True
        )
        st.markdown("<div style='font-size:9px;color:#555;letter-spacing:1px;margin-bottom:4px;'>IMPORT</div>",
                    unsafe_allow_html=True)
        up = st.file_uploader("Upload JSON", type=["json"], key="json_up",
                               label_visibility="collapsed")
        if up:
            try:
                raw = json.loads(up.read())
                for n in raw:
                    n.setdefault("nodeId", n.get("id",""))
                    n.setdefault("ftLabel","")
                    n.setdefault("fixedValue",None)
                    n.setdefault("targetValue",None)
                    n.setdefault("calculatedValue",None)
                    n.setdefault("parentIds",[])
                    n.setdefault("gate","OR")
                    n.setdefault("type","IF")
                set_nodes(raw, recalc=True)
                st.success(f"Imported {len(raw)} nodes")
                st.rerun()
            except Exception as e:
                st.error(f"Import failed: {e}")

        st.markdown("<div style='font-size:9px;color:#555;letter-spacing:1px;margin:8px 0 4px;'>EXPORT</div>",
                    unsafe_allow_html=True)
        if nodes:
            st.download_button("↓ JSON", export_json(nodes),
                               file_name=f"fta_{now_str()}.json", mime="application/json",
                               use_container_width=True)
            cypher_data = export_cypher(nodes)
            st.download_button("↓ Cypher (Neo4j)", cypher_data,
                               file_name=f"fta_{now_str()}.cypher", mime="text/plain",
                               use_container_width=True)
            xl = export_excel(nodes)
            if xl:
                st.download_button("↓ Excel", xl,
                                   file_name=f"fta_{now_str()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

    # ── HAZARD MANAGEMENT ────────────────────────────────────────────────
    with st.expander("⚠ HAZARDS", expanded=True):
        hname = st.text_input("Hazard name", placeholder="Loss of Propulsion", key="haz_name")
        htgt  = st.text_input("Target probability", placeholder="1e-7", key="haz_tgt")
        hgate = st.radio("Gate", ["OR","AND"], horizontal=True, key="haz_gate")
        if st.button("＋ ADD HAZARD", use_container_width=True, type="primary"):
            if hname.strip():
                try:    tv = float(htgt) if htgt.strip() else 1e-7
                except: tv = 1e-7
                nid = str(uuid.uuid4())[:7]
                new_h = {"id":nid,"nodeId":f"HAZ-{len(hazards)+1}","ftLabel":"",
                         "name":hname.strip(),"type":"HAZARD","gate":hgate,
                         "fixedValue":None,"targetValue":tv,"calculatedValue":tv,"parentIds":[]}
                st.session_state.tree_state["focus_id"] = nid
                set_nodes(nodes + [new_h])
                st.rerun()

        if hazards:
            st.markdown("---")
            for h in hazards:
                c = "#ff4d4d"
                tv = h.get("targetValue")
                cv = h.get("calculatedValue")
                st.markdown(
                    f"<div style='background:#1a0a0a;border:1px solid {c}33;border-radius:5px;"
                    f"padding:5px 8px;margin-bottom:3px;'>"
                    f"<div style='font-size:10px;color:{c};font-weight:700;'>{esc(h['name'])}</div>"
                    f"<div style='font-size:9px;color:#555;font-family:monospace;'>"
                    f"target: {fmt(tv)} · calc: {fmt(cv)}</div></div>",
                    unsafe_allow_html=True)

    # ── NODE MANAGEMENT ──────────────────────────────────────────────────
    tab_add, tab_edit, tab_shared = st.tabs(["➕ ADD", "✏️ EDIT", "⚡ SHARED"])

    with tab_add:
        nodes = st.session_state.nodes
        by_id = {n["id"]: n for n in nodes}

        node_name = st.text_input("Node name", placeholder="Valve fails to close", key="add_name")

        nc1, nc2 = st.columns([2, 1])
        with nc1:
            valid_prefixes = ["IF","FF","SF","GROUP","HAZ","OTHER"]
            add_prefix = st.selectbox("ID Prefix", valid_prefixes, key="add_prefix",
                                      label_visibility="visible")
        with nc2:
            add_num = st.text_input("Number", placeholder="196", key="add_num",
                                    label_visibility="visible")
        add_ft = st.text_input("FT Label (optional)", placeholder="FT-46", key="add_ft",
                               label_visibility="visible")
        cid_clean = f"{add_prefix}-{add_num.strip()}" if add_num.strip() else ""
        ft_clean  = re.sub(r'[^A-Za-z0-9\-_\.]','', add_ft.strip()) if add_ft.strip() else ""

        # Duplicate detection
        name_matches = [n for n in nodes if n["name"].lower() == node_name.strip().lower()] if node_name.strip() else []
        id_matches   = [n for n in nodes if (n.get("nodeId","") or "").upper() == cid_clean.upper()] if cid_clean else []

        def render_match_card(matches, match_type):
            for ex in matches:
                ex_color = LEVEL_COLORS.get(ex["type"],"#888")
                ex_val   = fmt(ex.get("calculatedValue"))
                ex_pnames = " · ".join(by_id[p]["name"] for p in (ex.get("parentIds") or []) if p in by_id) or "—"
                ex_nid    = ex.get("nodeId", ex["id"])
                def find_hazards_of(node_id, depth=0):
                    if depth > 8: return []
                    n = by_id.get(node_id)
                    if not n: return []
                    if n["type"] == "HAZARD": return [n["name"]]
                    result = []
                    for pid in (n.get("parentIds") or []):
                        result.extend(find_hazards_of(pid, depth+1))
                    return list(dict.fromkeys(result))
                ex_fts    = find_hazards_of(ex["id"])
                ex_ft_str = " · ".join(ex_fts) if ex_fts else "—"
                st.markdown(f"""
                <div style="background:#080f1a;border:1.5px solid #4fc3f7;border-radius:7px;padding:9px 12px;margin:3px 0 4px 0;">
                  <div style="font-size:8px;color:#4fc3f7;font-weight:700;letter-spacing:1px;margin-bottom:4px;">⚠ NODE ALREADY EXISTS</div>
                  <div style="font-size:10px;color:{ex_color};font-weight:700;">{esc(ex['name'])}</div>
                  <div style="font-size:9px;color:#777;">Value: <span style="font-family:monospace;color:{ex_color};">{ex_val}</span> · In FT: {ex_ft_str}</div>
                  <div style="font-size:9px;color:#4fc3f7;margin-top:4px;">Same failure — select parent(s) below then click Place.</div>
                </div>""", unsafe_allow_html=True)
                if st.button(f"◈ Place under selected parent(s)", key=f"place_{match_type}_{ex['id']}", use_container_width=True, type="primary"):
                    cur_sel = sel_pids if sel_pids else []
                    if not cur_sel:
                        st.error("Select at least one parent above first")
                    else:
                        updated = []
                        for n in nodes:
                            if n["id"] == ex["id"]:
                                n = dict(n)
                                ep = list(n.get("parentIds") or [])
                                ep += [p for p in cur_sel if p not in ep]
                                n["parentIds"] = ep
                            updated.append(n)
                        st.session_state.tree_state["focus_id"] = ex["id"]
                        st.session_state.nodes_since_calc += 1
                        set_nodes(updated)
                        st.success(f"✓ Placed under new parent(s).")
                        st.rerun()

        # Deduplicate — same node can appear in both name_matches and id_matches
        _seen_match_ids = set()
        _all_matches = []
        for _m in (name_matches + id_matches):
            if _m["id"] not in _seen_match_ids:
                _seen_match_ids.add(_m["id"])
                _all_matches.append(_m)
        if _all_matches: render_match_card(_all_matches, "combined")

        if not cid_clean:
            st.markdown("<div style='font-size:9px;color:#e94560;margin:2px 0 4px 0;'>★ Node ID required</div>", unsafe_allow_html=True)
        elif len(cid_clean) >= 2 and not id_matches:
            st.markdown(f"<div style='background:#0a1a0a;border:1px solid #4caf7d;border-radius:5px;padding:4px 8px;margin:2px 0 4px 0;'><span style='font-size:9px;color:#4caf7d;font-weight:700;'>✓ {cid_clean} available</span></div>", unsafe_allow_html=True)

        parent_opts = {f"[{n['type']}] {n.get('nodeId',n['id'])} — {n['name']}": n["id"]
                       for n in nodes if n["type"] in VALID_PARENT_TYPES
                       or (n["type"] in VALID_CHILD_TYPES and not n.get("parentIds"))}
        sel_labels = st.multiselect("Parent Node(s)", list(parent_opts.keys()), key="add_par")
        sel_pids   = [parent_opts[l] for l in sel_labels]

        node_type = st.selectbox("Type", VALID_CHILD_TYPES, key="add_type")
        gate      = st.radio("Gate", ["OR","AND"], horizontal=True, key="add_gate")
        use_fixed = st.checkbox("📌 Pin to fixed value", key="add_use_fixed")
        fixed_val_input = None
        if use_fixed:
            fixed_val_input = st.text_input("Fixed Value", placeholder="1.67e-9", key="add_fixed_val")

        add_btn = st.button("✅ ADD NODE", use_container_width=True, type="primary", disabled=not cid_clean)
        if add_btn:
            if not cid_clean:
                st.error("Node ID is required")
            elif not node_name.strip():
                st.error("Enter a node name")
            else:
                fv = None
                if use_fixed and fixed_val_input:
                    try:    fv = float(fixed_val_input)
                    except: st.error("Fixed value must be a number"); fv = None
                display_name = node_name.strip()
                if ft_clean:
                    display_name = f"[{ft_clean}] {display_name}"
                nid = str(uuid.uuid4())[:7]
                new_node = {"id":nid,"nodeId":cid_clean,"ftLabel":ft_clean or "","name":display_name,
                            "type":node_type,"gate":gate,"fixedValue":fv,"targetValue":None,
                            "calculatedValue":fv,"parentIds":sel_pids}
                st.session_state.tree_state["focus_id"] = nid
                st.session_state.nodes_since_calc += 1
                pnn = st.session_state.get("pending_node_names",[])
                pnn.append(f"{cid_clean} {display_name[:30]}")
                st.session_state.pending_node_names = pnn[-20:]
                set_nodes(nodes + [new_node])
                st.rerun()

        st.markdown("---")
        del_opts = {f"[{n['type']}] {n.get('nodeId',n['id'])} — {n['name']}": n["id"]
                    for n in nodes if n["type"] != "HAZARD"}
        if del_opts:
            dl = st.selectbox("Delete Node", ["— select —"] + list(del_opts.keys()), key="del_sel")
            if dl != "— select —":
                del_id = del_opts[dl]
                if st.button("🗑 DELETE NODE", use_container_width=True):
                    pre_existing_roots = {n["id"] for n in nodes if n["type"] != "HAZARD" and not n.get("parentIds")}
                    temp_nodes = [dict(n) for n in nodes if n["id"] != del_id]
                    for n in temp_nodes:
                        if del_id in (n.get("parentIds") or []):
                            n["parentIds"] = [p for p in n["parentIds"] if p != del_id]
                    changed = True
                    while changed:
                        changed = False
                        orphan_ids = {n["id"] for n in temp_nodes if n["type"] != "HAZARD" and not n.get("parentIds") and n["id"] not in pre_existing_roots}
                        if orphan_ids:
                            temp_nodes = [n for n in temp_nodes if n["id"] not in orphan_ids]
                            for n in temp_nodes:
                                before = len(n.get("parentIds") or [])
                                n["parentIds"] = [p for p in (n.get("parentIds") or []) if p not in orphan_ids]
                                if len(n.get("parentIds") or []) != before: changed = True
                    st.session_state.nodes_since_calc += 1
                    set_nodes(temp_nodes)
                    st.rerun()

        st.markdown("---")
        if st.button("🗑 CLEAR ALL", use_container_width=True):
            set_nodes([])
            st.session_state.selected_id = None
            st.session_state.nodes_since_calc = 0
            st.rerun(scope="app")

    with tab_edit:
        nodes = st.session_state.nodes
        by_id = {n["id"]: n for n in nodes}
        if not nodes:
            st.markdown("<div style='color:#555;font-size:11px;'>No nodes yet.</div>", unsafe_allow_html=True)
        else:
            edit_opts  = {f"[{n['type']}] {n.get('nodeId',n['id'])} — {n['name']}": n["id"] for n in nodes}
            # Auto-select node if triggered from canvas click → Edit button
            _et = st.session_state.pop("_sidebar_edit_target", None)
            _et_label = None
            if _et:
                _et_label = next((k for k,v in edit_opts.items() if v == _et), None)
            _default_edit_idx = 0
            if _et_label and _et_label in (["— select —"] + list(edit_opts.keys())):
                _default_edit_idx = (["— select —"] + list(edit_opts.keys())).index(_et_label)
            edit_label = st.selectbox("Select node to edit", ["— select —"] + list(edit_opts.keys()),
                                      index=_default_edit_idx, key="edit_sel")
            if edit_label != "— select —":
                eid = edit_opts[edit_label]
                en  = next((n for n in nodes if n["id"] == eid), None)
                prev_eid = st.session_state.get("_edit_prev_eid")
                if prev_eid != eid:
                    for k in ["en_name","en_nid_prefix","en_nid_num","en_ft","en_gate","en_type","en_par","en_tgt","en_use_fix","en_fv"]:
                        st.session_state.pop(k, None)
                    st.session_state["_edit_prev_eid"] = eid
                if en:
                    color = LEVEL_COLORS.get(en["type"], "#888")
                    st.markdown(f"""
                    <div style="background:#141414;border:2px solid {color};border-radius:8px;padding:8px 12px;margin-bottom:8px;">
                      <div style="font-size:8px;color:#888;letter-spacing:2px;">EDITING</div>
                      <div style="font-weight:700;color:{color};font-size:13px;">{esc(en['name'])}</div>
                      <div style="font-size:11px;color:{color};font-family:monospace;">{fmt(en.get('calculatedValue'))}</div>
                    </div>""", unsafe_allow_html=True)

                    new_name = st.text_input("Name", value=en["name"], key="en_name")
                    cur_nid = en.get("nodeId", en["id"])
                    import re as _re
                    nid_match  = _re.match(r'^([A-Za-z]+)-(.+)$', cur_nid)
                    cur_prefix = nid_match.group(1).upper() if nid_match else "IF"
                    cur_num    = nid_match.group(2) if nid_match else cur_nid
                    ec1, ec2, ec3 = st.columns([1, 2, 1])
                    with ec1:
                        vp = ["IF","FF","SF","GROUP","HAZ","OTHER"]
                        pi = vp.index(cur_prefix) if cur_prefix in vp else 0
                        ep = st.selectbox("Prefix", vp, index=pi, key="en_nid_prefix", label_visibility="collapsed")
                    with ec2:
                        en_num = st.text_input("Number", value=cur_num, key="en_nid_num", label_visibility="collapsed")
                    with ec3:
                        en_ft = st.text_input("FT", value=en.get("ftLabel",""), key="en_ft", label_visibility="collapsed")
                    new_nid = f"{ep}-{en_num.strip()}" if en_num.strip() else cur_nid
                    new_ft  = re.sub(r'[^A-Za-z0-9\-_\.]','', en_ft.strip()) if en_ft.strip() else ""
                    valid_types = ["HAZARD"] + VALID_CHILD_TYPES if en["type"] == "HAZARD" else VALID_CHILD_TYPES
                    ti = valid_types.index(en["type"]) if en["type"] in valid_types else 0
                    new_type = st.selectbox("Type", valid_types, index=ti, key="en_type")
                    new_gate = st.radio("Gate", ["OR","AND"], index=0 if en.get("gate","OR")=="OR" else 1, horizontal=True, key="en_gate")

                    if en["type"] != "HAZARD":
                        par_opts = {f"[{n['type']}] {n.get('nodeId',n['id'])} — {n['name']}": n["id"]
                                    for n in nodes if n["id"] != eid and n["type"] in VALID_PARENT_TYPES}
                        cur_par_labels = [f"[{by_id[p]['type']}] {by_id[p].get('nodeId',p)} — {by_id[p]['name']}"
                                          for p in (en.get("parentIds") or []) if p in by_id]
                        new_pars = st.multiselect("Parents", list(par_opts.keys()), default=cur_par_labels, key="en_par")
                        new_par_ids = [par_opts[l] for l in new_pars]
                    else:
                        new_par_ids = []

                    if en["type"] == "HAZARD":
                        tv_str = str(en.get("targetValue","")) if en.get("targetValue") is not None else ""
                        new_tgt_str = st.text_input("Target Probability", value=tv_str, key="en_tgt")
                        try:    new_tgt = float(new_tgt_str) if new_tgt_str.strip() else None
                        except: new_tgt = en.get("targetValue")
                    else:
                        new_tgt = en.get("targetValue")

                    use_fix = st.checkbox("📌 Pin fixed value", value=en.get("fixedValue") is not None, key="en_use_fix")
                    new_fv = en.get("fixedValue")
                    if use_fix:
                        fv_str = str(en.get("fixedValue","")) if en.get("fixedValue") is not None else ""
                        fv_inp = st.text_input("Fixed Value", value=fv_str, key="en_fv")
                        try:    new_fv = float(fv_inp) if fv_inp.strip() else None
                        except: new_fv = en.get("fixedValue")
                    else:
                        new_fv = None

                    if st.button("💾 SAVE CHANGES", use_container_width=True, type="primary"):
                        dn = new_name.strip()
                        if new_ft: dn = f"[{new_ft}] {dn}" if not dn.startswith(f"[{new_ft}]") else dn
                        updated = []
                        for n in nodes:
                            if n["id"] == eid:
                                n = dict(n)
                                n["name"] = dn; n["nodeId"] = new_nid; n["ftLabel"] = new_ft
                                n["type"] = new_type; n["gate"] = new_gate
                                n["fixedValue"] = new_fv; n["targetValue"] = new_tgt
                                if en["type"] != "HAZARD": n["parentIds"] = new_par_ids
                            updated.append(n)
                        st.session_state.tree_state["focus_id"] = eid
                        st.session_state.nodes_since_calc += 1
                        set_nodes(updated)
                        st.success("✓ Saved")
                        st.rerun()

    with tab_shared:
        nodes  = st.session_state.nodes
        by_id  = {n["id"]: n for n in nodes}
        if not nodes:
            st.markdown("<div style='color:#555;font-size:11px;'>No nodes yet.</div>", unsafe_allow_html=True)
        else:
            from collections import defaultdict as _dd2
            nid_map = _dd2(list)
            for n in nodes:
                nid = (n.get("nodeId") or "").strip()
                if nid: nid_map[nid].append(n)
            dup_groups  = {nid: grp for nid, grp in nid_map.items() if len(grp) > 1}
            link_shared = [n for n in nodes if len(n.get("parentIds") or []) > 1]
            c1, c2 = st.columns(2)
            with c1:
                col = "#f5c518" if dup_groups else "#333"
                st.markdown(f'<div style="background:#141414;border:1px solid {col}44;border-radius:6px;padding:7px;text-align:center;"><div style="font-size:8px;color:#555;">DUPLICATE</div><div style="font-size:22px;font-weight:700;color:{col};">{len(dup_groups)}</div></div>', unsafe_allow_html=True)
            with c2:
                col2 = "#4fc3f7" if link_shared else "#333"
                st.markdown(f'<div style="background:#141414;border:1px solid {col2}44;border-radius:6px;padding:7px;text-align:center;"><div style="font-size:8px;color:#555;">LINK-SHARED</div><div style="font-size:22px;font-weight:700;color:{col2};">{len(link_shared)}</div></div>', unsafe_allow_html=True)
            for nid_lbl, grp in sorted(dup_groups.items()):
                gc = LEVEL_COLORS.get(grp[0]["type"],"#888")
                pinned_vals = [n["fixedValue"] for n in grp if n.get("fixedValue") is not None]
                pin_note = f" · MAX={fmt(max(pinned_vals))}" if pinned_vals else ""
                st.markdown(
                    f"<div style='background:#141414;border:1px solid {gc}44;border-radius:6px;"
                    f"padding:6px 10px;margin:3px 0;'>"
                    f"<span style='color:{gc};font-weight:700;font-size:10px;'>◈ {esc(nid_lbl)}</span>"
                    f"<span style='font-size:9px;color:#555;'> · {len(grp)} instances{pin_note}</span></div>",
                    unsafe_allow_html=True)

    # ── CALCULATE button ─────────────────────────────────────────────────
    st.markdown("---")
    pending = st.session_state.nodes_since_calc
    calc_label = f"⚡ CALCULATE" + (f" ({pending} pending)" if pending else "")
    if st.button(calc_label, use_container_width=True, type="primary",
                 key="calc_btn"):
        set_nodes(st.session_state.nodes, recalc=True)
        st.rerun(scope="app")

    # ── SNAPSHOT ─────────────────────────────────────────────────────────
    if configured:
        st.markdown("---")
        if st.button("📸 Snapshot", use_container_width=True, key="snap_btn",
                     help="Save a timestamped copy of current state"):
            snap_name = f"snapshot_{now_str()}.json"
            save_current(st.session_state.nodes, snap_name, f"Snapshot saved: {snap_name}")
            st.session_state.file_list = list_gist_files(GITHUB_TOKEN, GIST_ID)
            st.success(f"Snapshot: {snap_name}")
        snaps = [f for f in st.session_state.file_list if is_snap(f)]
        if snaps:
            del_snap = st.selectbox("Delete snapshot", ["— keep —"] + snaps, key="del_snap")
            if del_snap != "— keep —" and st.button("🗑 Delete Snapshot", key="del_snap_btn"):
                del_gist_file(GITHUB_TOKEN, GIST_ID, del_snap)
                st.session_state.file_list = list_gist_files(GITHUB_TOKEN, GIST_ID)
                st.rerun()

# Handle canvas messages (positions + multisel + node edits)
cmp_html = """
<script>
window.addEventListener("message",function(e){
  try{
    const d=typeof e.data==="string"?JSON.parse(e.data):e.data;
    if(!d) return;
    if(d.type==="fta_pos"){
      window.parent.postMessage(JSON.stringify({type:"streamlit:setComponentValue",value:{type:"fta_pos",data:d.data}}),"*");
    }
    if(d.type==="fta_multisel"){
      window.parent.postMessage(JSON.stringify({type:"streamlit:setComponentValue",value:{type:"fta_multisel",data:d.data}}),"*");
    }
    if(d.type==="fta_selnode"){
      window.parent.postMessage(JSON.stringify({type:"streamlit:setComponentValue",value:{type:"fta_selnode",data:d.data}}),"*");
    }
    if(d.type==="fta_edit_node"){
      window.parent.postMessage(JSON.stringify({type:"streamlit:setComponentValue",value:{type:"fta_edit_node",data:d.data}}),"*");
    }
    if(d.type==="fta_delete_node"){
      window.parent.postMessage(JSON.stringify({type:"streamlit:setComponentValue",value:{type:"fta_delete_node",data:d.data}}),"*");
    }
  }catch(err){}
});
</script>"""
msg_val = components.html(cmp_html, height=0)
if msg_val and isinstance(msg_val, dict):
    if msg_val.get("type") == "fta_pos":
        new_pos = msg_val.get("data", {})
        if new_pos:
            st.session_state["_pending_positions"] = {**st.session_state.get("_pending_positions",{}), **new_pos}
            st.session_state.tree_state["positions"] = st.session_state["_pending_positions"]
    elif msg_val.get("type") == "fta_multisel":
        new_sel = [item["id"] for item in msg_val.get("data",[]) if "id" in item]
        if new_sel != st.session_state.get("multisel_ids",[]):
            st.session_state["multisel_ids"] = new_sel
            st.rerun(scope="fragment")
    elif msg_val.get("type") == "fta_selnode":
        sn_data = msg_val.get("data")  # None = deselect
        if sn_data != st.session_state.get("selected_node"):
            st.session_state["selected_node"] = sn_data
            st.rerun(scope="fragment")
    elif msg_val.get("type") == "fta_edit_node":
        ed = msg_val.get("data", {})
        eid = ed.get("id")
        if eid:
            _nodes = st.session_state.nodes
            _changed = False
            for _n in _nodes:
                if _n["id"] == eid:
                    if ed.get("name"):      _n["name"]       = ed["name"];      _changed = True
                    if ed.get("gate"):      _n["gate"]       = ed["gate"];      _changed = True
                    if ed.get("nodeId"):    _n["nodeId"]     = ed["nodeId"];    _changed = True
                    if "ftLabel" in ed:     _n["ftLabel"]    = ed.get("ftLabel",""); _changed = True
                    fv = ed.get("fixedVal")
                    _n["fixedValue"] = fv if fv is not None else None
                    _n["calculatedValue"] = fv if fv is not None else _n.get("calculatedValue")
                    tv = ed.get("targetVal")
                    if tv is not None: _n["targetValue"] = tv
                    _changed = True
                    break
            if _changed:
                st.session_state.nodes_since_calc += 1
                set_nodes(_nodes)
                st.rerun(scope="fragment")
    elif msg_val.get("type") == "fta_delete_node":
        del_id = msg_val.get("data", {}).get("id")
        if del_id:
            _nodes = st.session_state.nodes
            _pre_roots = {n["id"] for n in _nodes if n["type"] != "HAZARD" and not n.get("parentIds")}
            _tmp = [dict(n) for n in _nodes if n["id"] != del_id]
            for _n in _tmp:
                if del_id in (_n.get("parentIds") or []):
                    _n["parentIds"] = [p for p in _n["parentIds"] if p != del_id]
            _chg = True
            while _chg:
                _chg = False
                _orphans = {n["id"] for n in _tmp if n["type"] != "HAZARD" and not n.get("parentIds") and n["id"] not in _pre_roots}
                if _orphans:
                    _tmp = [n for n in _tmp if n["id"] not in _orphans]
                    for _n in _tmp:
                        _before = len(_n.get("parentIds") or [])
                        _n["parentIds"] = [p for p in (_n.get("parentIds") or []) if p not in _orphans]
                        if len(_n.get("parentIds") or []) != _before: _chg = True
            st.session_state["selected_node"] = None
            st.session_state.nodes_since_calc += 1
            set_nodes(_tmp)
            st.rerun()

with st.sidebar:
    render_sidebar()

# ── TAB 1: TREE ───────────────────────────────────────────────────────────
with tab_tree:
    if not nodes:
        st.markdown("""
        <div style='text-align:center;color:#333;margin-top:60px;'>
          <div style='font-size:32px;'>⚠</div>
          <div style='font-size:14px;color:#555;margin-top:8px;'>No nodes yet</div>
          <div style='font-size:10px;color:#333;margin-top:4px;'>Add a HAZARD node in the sidebar to begin</div>
        </div>""", unsafe_allow_html=True)
    else:
        filt_opts  = {"ALL": None} | {h["name"]: h["id"] for h in hazards}
        filt_label = st.selectbox("Filter by hazard", list(filt_opts.keys()),
                                  key="tree_filter_sel", label_visibility="collapsed")
        filt_id    = filt_opts[filt_label]
        tree_html  = build_html_tree(nodes, filter_hazard_id=filt_id,
                                     tree_state=st.session_state.tree_state)
        if tree_html:
            components.html(tree_html, height=780, scrolling=False)
        else:
            st.info("No nodes visible for this filter.")

# ── TAB 2: VERIFY (NEW — PFTA integration) ───────────────────────────────
with tab_verify:
    render_pfta_verification(nodes)

# ── TAB 3: HIERARCHY ─────────────────────────────────────────────────────
with tab_hier:
    if not nodes:
        st.markdown("<div style='color:#333;text-align:center;margin-top:40px;'>No nodes yet</div>",
                    unsafe_allow_html=True)
    else:
        fh_opts  = {"ALL hazards": None} | {h["name"]: h["id"] for h in hazards}
        fh_label = st.selectbox("Filter", list(fh_opts.keys()), key="hier_filter",
                                label_visibility="collapsed")
        fh_id    = fh_opts[fh_label]
        rows     = build_hierarchy_rows(nodes, filter_hazard_id=fh_id)

        for r in rows:
            node  = r["node"]
            depth = r["depth"]
            is_ref = r["ref"]
            color  = LEVEL_COLORS.get(node["type"], "#7e57c2")
            val    = fmt(node.get("calculatedValue"))
            indent = depth * 20
            val_color  = "#e94560" if node.get("fixedValue") is not None else color
            val_display = val + (" 📌" if node.get("fixedValue") is not None else "")
            tgt_str = fmt(node.get("targetValue")) if node.get("targetValue") else ""
            node_id = node.get("nodeId", node["id"])
            is_shared = len(node.get("parentIds") or []) > 1
            pnames = esc(" · ".join(by_id[p]["name"] for p in (node.get("parentIds") or []) if p in by_id) or "—")
            cnames = esc(" · ".join(n["name"] for n in nodes if node["id"] in (n.get("parentIds") or [])) or "—")
            shape_style = "border-radius:50px;" if node["type"] == "GROUP" else "border-radius:6px;"
            pin_border  = "border:2px solid #e94560;" if node.get("fixedValue") is not None else f"border:1px solid {color}33;"
            ref_style   = "opacity:0.55;border-style:dashed;" if is_ref else ""
            shr_badge   = '<span style="background:#f5c518;color:#111;font-size:7px;padding:0 3px;border-radius:3px;margin-left:5px;font-weight:700;">SHR</span>' if is_shared else ""
            ref_badge   = '<span style="background:#4444aa;color:#fff;font-size:7px;padding:0 3px;border-radius:3px;margin-left:3px;">REF</span>' if is_ref else ""
            tgt_row     = f'<div style="font-size:8px;color:#555;font-family:monospace;">target: {tgt_str}</div>' if tgt_str else ""
            html_block = (
                f'<div style="margin-left:{indent}px;background:#141414;{pin_border}{shape_style}{ref_style}'
                f'padding:7px 12px;margin-bottom:3px;">'
                f'<div style="display:grid;grid-template-columns:2fr 0.6fr 0.6fr 1.2fr 1.8fr;gap:8px;align-items:center;">'
                f'<div><span style="font-size:11px;font-weight:700;color:#ddd;">{esc(node["name"])}</span>{shr_badge}{ref_badge}</div>'
                f'<div><div style="font-size:7px;color:#444;letter-spacing:1px;">NODE ID</div>'
                f'<div style="font-size:10px;color:{color};font-weight:700;font-family:monospace;">{esc(node_id)}</div></div>'
                f'<div><div style="font-size:7px;color:#444;letter-spacing:1px;">TYPE</div>'
                f'<div style="font-size:10px;color:{color};font-weight:700;">{esc(node["type"])}</div></div>'
                f'<div><div style="font-size:7px;color:#444;letter-spacing:1px;">CALC VALUE</div>'
                f'<div style="font-size:11px;color:{val_color};font-weight:700;font-family:monospace;">{val_display}</div>'
                f'{tgt_row}</div>'
                f'<div><div style="font-size:9px;color:#555;">&#8593; {pnames}</div>'
                f'<div style="font-size:9px;color:#444;">&#8595; {cnames}</div></div>'
                f'</div></div>'
            )
            st.markdown(html_block, unsafe_allow_html=True)

        st.markdown("---")
        cols = st.columns(6)
        counts = [(lvl, len(by_level[lvl])) for lvl in DISPLAY_ORDER]
        counts += [("TOTAL", len(nodes)), ("📌 PINNED", sum(1 for n in nodes if n.get("fixedValue") is not None))]
        for i,(lvl,cnt) in enumerate(counts):
            with cols[i%6]:
                c = "#e94560" if lvl == "📌 PINNED" else LEVEL_COLORS.get(lvl,"#e94560")
                st.markdown(f"""<div style="background:#141414;border:1px solid {c}44;border-radius:5px;padding:8px;text-align:center;">
                  <div style="font-size:8px;color:#555;letter-spacing:2px;">{lvl}</div>
                  <div style="font-size:18px;font-weight:700;color:{c};">{cnt}</div>
                </div>""", unsafe_allow_html=True)

# ── TAB 4: DATA ───────────────────────────────────────────────────────────
with tab_data:
    if not nodes:
        st.markdown("<div style='color:#333;text-align:center;margin-top:40px;'>No nodes yet</div>",
                    unsafe_allow_html=True)
    else:
        show_by_level = {lvl: [n for n in nodes if n["type"] == lvl] for lvl in DISPLAY_ORDER}
        for level in DISPLAY_ORDER:
            lvl_nodes = show_by_level[level]
            if not lvl_nodes: continue
            color = LEVEL_COLORS.get(level, "#7e57c2")
            st.markdown(f"<div style='font-size:9px;letter-spacing:3px;color:{color};border-bottom:1px solid {color}33;padding-bottom:3px;margin:10px 0 5px;'>{level} — {len(lvl_nodes)} nodes</div>", unsafe_allow_html=True)
            for node in lvl_nodes:
                color     = LEVEL_COLORS.get(node["type"], "#7e57c2")
                gc        = "#4fc3f7" if node["gate"] == "OR" else "#ffb74d"
                val       = fmt(node.get("calculatedValue"))
                node_id   = node.get("nodeId", node["id"])
                ft_lbl    = node.get("ftLabel","")
                pnames    = esc(" · ".join(by_id[p]["name"] for p in (node.get("parentIds") or []) if p in by_id) or "—")
                cnames    = esc(" · ".join(n["name"] for n in nodes if node["id"] in (n.get("parentIds") or [])) or "—")
                is_shared = len(node.get("parentIds") or []) > 1
                is_pinned = node.get("fixedValue") is not None
                val_color = "#e94560" if is_pinned else color
                tgt_str   = fmt(node.get("targetValue")) if node.get("targetValue") else ""
                shape_style = "border-radius:50px;" if node["type"] == "GROUP" else "border-radius:6px;"
                pin_border  = "border:2px solid #e94560;" if is_pinned else f"border:2px solid {color}44;"
                badges = f'<code style="background:#1a1a2e;color:{color};font-size:9px;padding:1px 5px;border-radius:3px;font-weight:700;">{esc(node_id)}</code>'
                if ft_lbl:
                    badges += f' <code style="background:#1a1a2e;color:#7e57c2;font-size:9px;padding:1px 5px;border-radius:3px;">{esc(ft_lbl)}</code>'
                if is_shared:
                    badges += ' <span style="background:#f5c518;color:#111;font-size:7px;padding:1px 4px;border-radius:3px;font-weight:700;">SHARED</span>'
                if is_pinned:
                    badges += f' <span style="background:#e9456022;color:#e94560;font-size:7px;padding:1px 4px;border-radius:3px;border:1px solid #e9456044;">&#128204; FIXED={fmt(node.get("fixedValue"))}</span>'
                tgt_row = f'<div style="font-size:8px;color:#555;font-family:monospace;">target: {tgt_str}</div>' if tgt_str else ""
                pin_icon = "&#128204;" if is_pinned else ""
                html_block = (
                    f'<div style="background:#141414;{pin_border}{shape_style}padding:9px 14px;margin-bottom:5px;">'
                    f'<div style="display:grid;grid-template-columns:2.5fr 0.7fr 0.7fr 1.5fr 2fr;gap:10px;align-items:start;">'
                    f'<div><div style="font-weight:700;font-size:11px;color:#ddd;margin-bottom:3px;">{esc(node["name"])}</div>'
                    f'<div>{badges}</div></div>'
                    f'<div><div style="font-size:7px;color:#444;">TYPE</div><div style="font-size:10px;color:{color};font-weight:700;">{node["type"]}</div></div>'
                    f'<div><div style="font-size:7px;color:#444;">GATE</div><div style="font-size:10px;color:{gc};font-weight:700;">{node["gate"]}</div></div>'
                    f'<div><div style="font-size:7px;color:#444;">CALC VALUE</div>'
                    f'<div style="font-size:11px;color:{val_color};font-weight:700;font-family:monospace;">{val}{pin_icon}</div>'
                    f'{tgt_row}</div>'
                    f'<div><div style="font-size:8px;color:#444;">&#8593; {pnames}</div>'
                    f'<div style="font-size:8px;color:#333;margin-top:2px;">&#8595; {cnames}</div></div>'
                    f'</div></div>'
                )
                st.markdown(html_block, unsafe_allow_html=True)

# ── TAB 5: SEARCH ─────────────────────────────────────────────────────────
with tab_search:
    if not nodes:
        st.markdown("<div style='color:#333;text-align:center;margin-top:40px;'>No nodes yet</div>", unsafe_allow_html=True)
    else:
        st.markdown("<div style='font-size:9px;color:#555;letter-spacing:2px;margin-bottom:10px;'>SEARCH ACROSS ALL NODES</div>", unsafe_allow_html=True)
        sq = st.text_input("Search", placeholder="e.g. IF-016, isolation, 1.25e-04, AND", key="search_q", label_visibility="collapsed")
        if sq.strip():
            lq = sq.strip().lower()
            matches = [n for n in nodes if (
                lq in n["name"].lower() or lq in n["type"].lower() or
                lq in n["gate"].lower() or lq in fmt(n.get("calculatedValue")).lower() or
                lq in (n.get("nodeId","")).lower() or lq in (n.get("ftLabel","")).lower()
            )]
            st.markdown(f"<div style='font-size:10px;color:#ff8c42;margin-bottom:8px;'>{len(matches)} result(s) for <b>\"{esc(sq)}\"</b></div>", unsafe_allow_html=True)
            for node in matches:
                color    = LEVEL_COLORS.get(node["type"], "#7e57c2")
                gc       = "#4fc3f7" if node["gate"] == "OR" else "#ffb74d"
                val      = fmt(node.get("calculatedValue"))
                node_id  = node.get("nodeId", node["id"])
                pnames   = esc(" · ".join(by_id[p]["name"] for p in (node.get("parentIds") or []) if p in by_id) or "—")
                cnames   = esc(" · ".join(n["name"] for n in nodes if node["id"] in (n.get("parentIds") or [])) or "—")
                is_shared = len(node.get("parentIds") or []) > 1
                is_pinned = node.get("fixedValue") is not None
                val_color = "#e94560" if is_pinned else color
                display_name = esc(node["name"])
                try:
                    idx = node["name"].lower().index(lq)
                    raw = node["name"]
                    display_name = (esc(raw[:idx]) +
                        f'<span style="background:#f5c518;color:#111;border-radius:2px;padding:0 2px;">{esc(raw[idx:idx+len(lq)])}</span>' +
                        esc(raw[idx+len(lq):]))
                except ValueError:
                    pass
                shr_badge = '<span style="background:#f5c518;color:#111;font-size:7px;padding:0 3px;border-radius:3px;margin-left:4px;">SHR</span>' if is_shared else ""
                pin_icon  = "&#128204;" if is_pinned else ""
                html_block = (
                    f'<div style="background:#141414;border:2px solid {color}44;border-radius:6px;padding:9px 14px;margin-bottom:5px;">'
                    f'<div style="display:grid;grid-template-columns:2.5fr 0.7fr 0.7fr 1.5fr 2fr;gap:10px;align-items:start;">'
                    f'<div><div style="font-weight:700;font-size:11px;color:#ddd;">{display_name}</div>'
                    f'<code style="font-size:9px;color:{color};">{esc(node_id)}</code>{shr_badge}</div>'
                    f'<div><div style="font-size:7px;color:#444;">TYPE</div><div style="font-size:10px;color:{color};font-weight:700;">{node["type"]}</div></div>'
                    f'<div><div style="font-size:7px;color:#444;">GATE</div><div style="font-size:10px;color:{gc};font-weight:700;">{node["gate"]}</div></div>'
                    f'<div><div style="font-size:7px;color:#444;">VALUE</div><div style="font-size:11px;color:{val_color};font-weight:700;font-family:monospace;">{val}{pin_icon}</div></div>'
                    f'<div><div style="font-size:8px;color:#444;">&#8593; {pnames}</div><div style="font-size:8px;color:#333;">&#8595; {cnames}</div></div>'
                    f'</div></div>'
                )
                st.markdown(html_block, unsafe_allow_html=True)
        else:
            for level in DISPLAY_ORDER:
                lvl_nodes = by_level[level]
                if not lvl_nodes: continue
                color = LEVEL_COLORS.get(level, "#7e57c2")
                st.markdown(f"<div style='font-size:9px;letter-spacing:3px;color:{color};border-bottom:1px solid {color}33;padding-bottom:3px;margin:10px 0 5px;'>{level} — {len(lvl_nodes)} nodes</div>", unsafe_allow_html=True)
                for node in lvl_nodes:
                    val   = fmt(node.get("calculatedValue"))
                    gc    = "#4fc3f7" if node["gate"] == "OR" else "#ffb74d"
                    pnames = " · ".join(by_id[p]["name"] for p in (node.get("parentIds") or []) if p in by_id) or "—"
                    is_shared = len(node.get("parentIds") or []) > 1
                    c = LEVEL_COLORS.get(node["type"], "#7e57c2")
                    st.markdown(f"""
                    <div style="background:#141414;border-left:3px solid {c};border-radius:0 5px 5px 0;
                                padding:5px 10px;margin-bottom:3px;
                                display:grid;grid-template-columns:2.5fr 0.7fr 0.7fr 1.5fr 2fr;gap:8px;align-items:center;">
                      <div style="font-size:10px;color:#ddd;">
                        {esc(node['name'])}{'<span style="background:#f5c518;color:#111;font-size:7px;padding:0 3px;border-radius:3px;margin-left:5px;">SHR</span>' if is_shared else ''}
                      </div>
                      <div style="font-size:9px;color:{c};">{node['type']}</div>
                      <div style="font-size:9px;color:{gc};">{node['gate']}</div>
                      <div style="font-size:10px;color:{c};font-family:monospace;font-weight:700;">{val}</div>
                      <div style="font-size:9px;color:#555;">{pnames}</div>
                    </div>""", unsafe_allow_html=True)
